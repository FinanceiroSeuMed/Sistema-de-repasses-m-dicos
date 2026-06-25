# -*- coding: utf-8 -*-
"""
Geradores do repasse do médico: Excel (arquivo interno) e PDF (enviado ao médico).

- Um arquivo por médico.
- Mostra o honorário RECALCULADO (não o valor bruto do convênio).
- Nome do arquivo: "Dr. {nome} {DD-MM}" (data do atendimento).
- Lembrete de preceptoria (semanal/mensal) no topo, quando houver.
"""

from __future__ import annotations

import io
import os
import re

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_INVALIDO = re.compile(r'[\\/:*?"<>|]+')


def _logo_path():
    """Caminho do logo da SeuMed (None se não existir)."""
    from django.conf import settings
    p = getattr(settings, 'LOGO_PATH', None)
    return str(p) if p and os.path.exists(str(p)) else None


def moeda(valor) -> str:
    if valor is None:
        return 'A definir'
    texto = f'{float(valor):,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
    return f'R$ {texto}'


def _data_bloco(bloco):
    datas = [p.data for p in bloco.procedimentos if p.data]
    return max(datas) if datas else None


def _titulo_clinica(bloco, unidade: str) -> str:
    """Topo do documento de repasse = só a CLÍNICA dos atendimentos daquele bloco
    (ex.: "Maringá - Matriz"), não o banner completo da MedPlus com todas as
    unidades. Cai para `unidade` se o bloco não trouxer clínica. (Diretoria 2026-06-23.)"""
    return (getattr(bloco, 'clinica', '') or '').strip() or unidade or 'Repasse Médico'


def nome_base(bloco) -> str:
    """Ex.: 'Dr. Heric Sakamoto 16-06' ou 'Dr. Carlos 08-05 - Paiçandu'."""
    nome = _INVALIDO.sub('', bloco.profissional).strip().rstrip('.').strip()
    data_ref = _data_bloco(bloco)
    dia = data_ref.strftime('%d-%m') if data_ref else 'sem-data'
    base = f'{nome} {dia}'.strip()
    clin = (getattr(bloco, 'clinica', '') or '').strip()
    if clin:
        curta = clin.split(' - ')[-1] if ' - ' in clin else clin
        base += f' - {_INVALIDO.sub("", curta).strip()}'
    return base


def _limpar_nome(nome) -> str:
    return _INVALIDO.sub('', str(nome or '')).strip().rstrip('.').strip()


def nome_base_anestesista(entry) -> str:
    """Ex.: 'Dra. Isabela Miwa Maeda 07-05 (Dr. Rodolpho...)'."""
    anest = _limpar_nome(entry.get('anestesista'))
    cirurgiao = _limpar_nome(entry.get('cirurgiao'))
    data = entry.get('data')
    dia = data.strftime('%d-%m') if data else 'sem-data'
    return f'{anest} {dia} ({cirurgiao})'.strip()


def pagaveis(bloco):
    """Só as linhas com honorário a receber (> 0). Linhas R$ 0,00 e 'a definir'
    não entram no Excel arquivado nem no PDF enviado ao médico."""
    return [p for p in bloco.procedimentos
            if p.status_calculo == 'calculado' and (p.honorario or 0) > 0]


def _total(bloco, linhas=None) -> float:
    linhas = pagaveis(bloco) if linhas is None else linhas
    return round(sum(p.honorario for p in linhas), 2)


def _ajuste_arredondamento(bloco, linhas=None) -> float:
    """Diferença entre o Total (soma em precisão cheia, arredondada — igual à
    OMIE) e a soma das linhas exibidas com 2 casas. Vira uma linha
    "Arredondamento" para o documento de conferência fechar centavo a centavo."""
    linhas = pagaveis(bloco) if linhas is None else linhas
    soma_exibida = round(sum(round(p.honorario or 0, 2) for p in linhas), 2)
    return round(_total(bloco, linhas) - soma_exibida, 2)


# --- Excel (arquivo interno) --------------------------------------------------

def _linhas_repasse(bloco):
    """Linhas da tabela do repasse (mesmo conteúdo no Excel e no PDF — idênticos).
    Devolve (cabeçalho, linhas_de_dados, total, n_pagaveis).

    Sem linha de "Arredondamento": a diferença de centavos (da soma em precisão
    cheia vs. as linhas exibidas com 2 casas) é EMBUTIDA em uma linha — preferindo
    uma de valor percentual/fracionado; se não houver, na última — para a soma
    exibida fechar com o Total, sem justificar a diferença. (Diretoria 2026-06-24.)"""
    linhas_pag = pagaveis(bloco)
    total = _total(bloco, linhas_pag)
    diff = _ajuste_arredondamento(bloco, linhas_pag)
    alvo = None
    if diff and linhas_pag:
        for p in linhas_pag:
            if round(p.honorario, 2) != round(p.honorario, 6):   # fração de centavo (percentual)
                alvo = p
                break
        if alvo is None:
            alvo = linhas_pag[-1]
    cab = ['Data', 'Paciente', 'Procedimento', 'Convênio', 'Qtd.', 'Honorário']
    dados = []
    for p in linhas_pag:
        val = round(p.honorario or 0, 2)
        if p is alvo:
            val = round(val + diff, 2)
        dados.append([p.data_texto, p.paciente, p.procedimento, p.convenio, p.quantidade, val])
    return cab, dados, total, len(linhas_pag)


def gerar_excel(bloco, unidade: str) -> bytes:
    """Excel de arquivamento — layout IDÊNTICO ao PDF enviado ao médico (mesmas
    colunas, cabeçalho e total), para que abrir o Excel e salvar em PDF dê o mesmo
    documento. SEM o aviso de preceptoria (esse fica só na revisão)."""
    data_ref = _data_bloco(bloco)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Repasse'

    negrito = Font(bold=True)
    preto14 = Font(bold=True, size=14, color='000000')   # filial em PRETO (diretoria 2026-06-25)
    cabec_fill = PatternFill('solid', fgColor='0B5FA5')
    cabec_font = Font(bold=True, color='FFFFFF')
    dir_ = Alignment(horizontal='right')

    # Identidade visual (logo) no topo; depois a FILIAL em preto, à esquerda.
    logo = _logo_path()
    topo = 1
    if logo:
        try:
            img = XLImage(logo)
            img.width, img.height = 176, 64   # 440x160 -> ~2.75:1
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 50
            topo = 2
        except Exception:
            pass
    ws.cell(topo, 1, _titulo_clinica(bloco, unidade)).font = preto14
    ws.cell(topo + 1, 1, 'Demonstrativo de Repasse Médico').font = negrito
    ws.cell(topo + 2, 1, f'Profissional: {bloco.profissional}')
    linha = topo + 3
    if getattr(bloco, 'razao_social', ''):
        ws.cell(linha, 1, f'Razão Social: {bloco.razao_social}'); linha += 1
    if data_ref:
        ws.cell(linha, 1, f'Data do atendimento: {data_ref.strftime("%d/%m/%Y")}'); linha += 1
    linha += 1

    cab, dados, total, _ = _linhas_repasse(bloco)
    for c, titulo_col in enumerate(cab, start=1):
        cel = ws.cell(linha, c, titulo_col)
        cel.fill = cabec_fill
        cel.font = cabec_font
    linha += 1

    for d in dados:
        ws.cell(linha, 1, d[0]); ws.cell(linha, 2, d[1]); ws.cell(linha, 3, d[2])
        ws.cell(linha, 4, d[3]); ws.cell(linha, 5, d[4])
        cel_h = ws.cell(linha, 6, d[5]); cel_h.number_format = 'R$ #,##0.00'
        linha += 1

    # Total à DIREITA: rótulo na coluna Qtd, valor logo abaixo da coluna Honorário.
    cel_rot = ws.cell(linha, 5, 'Total'); cel_rot.font = negrito; cel_rot.alignment = dir_
    cel_total = ws.cell(linha, 6, total)
    cel_total.number_format = 'R$ #,##0.00'
    cel_total.font = negrito
    cel_total.alignment = dir_

    larguras = [12, 26, 44, 16, 6, 14]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- PDF (enviado ao médico) --------------------------------------------------

def gerar_pdf(bloco, unidade: str) -> bytes:
    data_ref = _data_bloco(bloco)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title=f'Repasse {bloco.profissional}')
    estilos = getSampleStyleSheet()
    # Filial em PRETO, à esquerda, logo abaixo da identidade visual. (Diretoria 2026-06-25.)
    st_titulo = ParagraphStyle('t', parent=estilos['Title'], fontSize=15, spaceAfter=2,
                               alignment=0, textColor=colors.black)
    st_sub = ParagraphStyle('s', parent=estilos['Normal'], fontSize=9, textColor=colors.HexColor('#52606d'))
    st_info = ParagraphStyle('i', parent=estilos['Normal'], fontSize=10, spaceAfter=2)
    st_cel = ParagraphStyle('c', parent=estilos['Normal'], fontSize=8.5, leading=11)

    elems = []
    logo = _logo_path()
    if logo:
        try:
            im = RLImage(logo, width=46 * mm, height=46 / 2.75 * mm)   # mantém a proporção 440x160
            im.hAlign = 'LEFT'
            elems += [im, Spacer(1, 6)]
        except Exception:
            pass
    elems += [
        Paragraph(_titulo_clinica(bloco, unidade), st_titulo),
        Paragraph('Demonstrativo de Repasse Médico', st_sub),
        Spacer(1, 8),
        Paragraph(f'<b>Profissional:</b> {bloco.profissional}', st_info),
    ]
    if getattr(bloco, 'razao_social', ''):
        elems.append(Paragraph(f'<b>Razão Social:</b> {bloco.razao_social}', st_info))
    if data_ref:
        elems.append(Paragraph(f'<b>Data do atendimento:</b> {data_ref.strftime("%d/%m/%Y")}', st_info))
    # SEM aviso de preceptoria aqui — esse fica só na revisão. (Diretoria 2026-06-24.)
    elems.append(Spacer(1, 8))

    cab, linhas_dados, total, n_pag = _linhas_repasse(bloco)
    # Mesmas colunas do Excel (inclui Paciente) — documentos idênticos.
    dados = [cab]
    for d in linhas_dados:
        dados.append([d[0], Paragraph(d[1] or '', st_cel), Paragraph(d[2], st_cel),
                      d[3], str(d[4]), moeda(d[5])])
    # Total à DIREITA: rótulo na coluna Qtd, valor logo abaixo da coluna Honorário.
    dados.append(['', '', '', '', 'Total', moeda(total)])
    rt = len(dados) - 1

    tabela = Table(dados, colWidths=[20 * mm, 38 * mm, 54 * mm, 26 * mm, 12 * mm, 26 * mm], repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FA5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (4, 0), (5, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F7F9FC')]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, rt), (-1, rt), colors.HexColor('#EEF2F7')),
        ('FONTNAME', (4, rt), (5, rt), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elems.append(tabela)
    elems.append(Spacer(1, 10))
    elems.append(Paragraph(
        f'Total de {n_pag} procedimento(s) com repasse. '
        'Documento para conferência do profissional.', st_sub))

    doc.build(elems)
    return buf.getvalue()


def gerar_excel_anestesista(entry, unidade: str) -> bytes:
    """Excel (arquivo interno) do repasse do anestesista — mesmo padrão do médico."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Repasse'

    negrito = Font(bold=True)
    preto14 = Font(bold=True, size=14, color='000000')
    cabec_fill = PatternFill('solid', fgColor='0B5FA5')
    cabec_font = Font(bold=True, color='FFFFFF')

    logo = _logo_path()
    topo = 1
    if logo:
        try:
            img = XLImage(logo); img.width, img.height = 176, 64
            ws.add_image(img, 'A1'); ws.row_dimensions[1].height = 50; topo = 2
        except Exception:
            pass
    ws.cell(topo, 1, (entry.get('clinica') or '').strip() or unidade or 'Repasse de Anestesia').font = preto14
    ws.cell(topo + 1, 1, 'Demonstrativo de Repasse de Anestesia').font = negrito
    ws.cell(topo + 2, 1, f'Anestesista: {entry.get("anestesista")}')
    ws.cell(topo + 3, 1, f'Cirurgião: {entry.get("cirurgiao")}')
    linha = topo + 4
    data = entry.get('data')
    if data:
        ws.cell(linha, 1, f'Data: {data.strftime("%d/%m/%Y")}'); linha += 1
    linha += 1

    colunas = ['Data', 'Procedimento', 'Convênio', 'Qtd.']
    for c, titulo_col in enumerate(colunas, start=1):
        cel = ws.cell(linha, c, titulo_col)
        cel.fill = cabec_fill
        cel.font = cabec_font
    linha += 1

    for p in entry.get('cirurgias', []):
        ws.cell(linha, 1, p.data_texto)
        ws.cell(linha, 2, p.procedimento)
        ws.cell(linha, 3, p.convenio)
        ws.cell(linha, 4, p.quantidade)
        linha += 1

    cel_rot = ws.cell(linha, 1, 'Total:')
    cel_rot.font = negrito
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=3)
    cel_total = ws.cell(linha, 2, round(float(entry.get('valor') or 0), 2))
    cel_total.number_format = 'R$ #,##0.00'
    cel_total.alignment = Alignment(horizontal='left')
    cel_total.font = negrito

    for i, w in enumerate([12, 50, 18, 8], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_pdf_anestesista(entry, unidade: str) -> bytes:
    """PDF de repasse do anestesista: lista as cirurgias do dia (com o cirurgião)
    e o total fixo do anestesista."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title=f'Repasse anestesista {entry.get("anestesista")}')
    estilos = getSampleStyleSheet()
    st_titulo = ParagraphStyle('t', parent=estilos['Title'], fontSize=15, spaceAfter=2,
                               alignment=0, textColor=colors.black)
    st_sub = ParagraphStyle('s', parent=estilos['Normal'], fontSize=9, textColor=colors.HexColor('#52606d'))
    st_info = ParagraphStyle('i', parent=estilos['Normal'], fontSize=10, spaceAfter=2)
    st_cel = ParagraphStyle('c', parent=estilos['Normal'], fontSize=8.5, leading=11)

    data = entry.get('data')
    elems = []
    logo = _logo_path()
    if logo:
        try:
            im = RLImage(logo, width=46 * mm, height=46 / 2.75 * mm); im.hAlign = 'LEFT'
            elems += [im, Spacer(1, 6)]
        except Exception:
            pass
    elems += [
        Paragraph((entry.get('clinica') or '').strip() or unidade or 'Repasse de Anestesia', st_titulo),
        Paragraph('Demonstrativo de Repasse de Anestesia', st_sub),
        Spacer(1, 8),
        Paragraph(f'<b>Anestesista:</b> {entry.get("anestesista")}', st_info),
        Paragraph(f'<b>Cirurgião:</b> {entry.get("cirurgiao")}', st_info),
    ]
    if data:
        elems.append(Paragraph(f'<b>Data:</b> {data.strftime("%d/%m/%Y")}', st_info))
    elems.append(Spacer(1, 8))

    dados = [['Data', 'Procedimento', 'Convênio', 'Qtd.']]
    for p in entry.get('cirurgias', []):
        dados.append([p.data_texto, Paragraph(p.procedimento, st_cel), p.convenio, str(p.quantidade)])

    tabela = Table(dados, colWidths=[24 * mm, 96 * mm, 36 * mm, 14 * mm], repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FA5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FC')]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elems.append(tabela)
    elems.append(Spacer(1, 10))
    elems.append(Paragraph(
        f'<b>Total do repasse (anestesia):</b> {moeda(entry.get("valor"))}',
        ParagraphStyle('tot', parent=estilos['Normal'], fontSize=12)))
    elems.append(Spacer(1, 4))
    elems.append(Paragraph(
        f'{len(entry.get("cirurgias", []))} cirurgia(s). Documento para conferência.', st_sub))
    doc.build(elems)
    return buf.getvalue()
