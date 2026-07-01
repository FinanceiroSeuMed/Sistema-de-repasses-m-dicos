# -*- coding: utf-8 -*-
"""
Geradores dos arquivos de importação da OMIE (contas a pagar e a receber)
a partir de um relatório da MedPlus já processado (com honorários calculados).

Regras definidas pela diretoria:
- Conta corrente: "Omie.CASH".
- Vencimento: dia 10 do mês seguinte ao mês do atendimento.
- A pagar: uma linha por (médico × dia × clínica × classe), categoria da classe.
  Só entram honorários CALCULADOS (> 0). Linhas "a definir" ou sem categoria não
  entram (viram pendência para o usuário resolver na revisão).
- A receber: uma linha por (dia × clínica) — recebimento geral dos pacientes —
  com o valor bruto somado; Cliente = a associação (CLIENTE_RECEBER), filial no
  Departamento; categoria = fallback único.
"""

from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from datetime import date

from openpyxl import load_workbook

from . import medplus, regras

# Mapeamento classe -> categoria da OMIE (contas a pagar). Textos confirmados pela
# diretoria (2026-06-30) — têm de bater EXATO com as categorias cadastradas na OMIE.
CATEGORIA_POR_CLASSE = {
    medplus.CLASSE_CIRURGIA: 'Repasse Oftalmologistas - Cirurgia',
    medplus.CLASSE_EXAME: 'Repasse Oftalmologistas - Consulta',
    medplus.CLASSE_LAUDO: 'Repasse Laudos',
    medplus.CLASSE_PRECEPTORIA: 'Preceptoria',
}
CATEGORIA_ANESTESISTA = 'Repasse Anestesiologistas'

# Taxa de sala NÃO entra no a pagar (só no a receber).
CLASSES_FORA_DO_PAGAR = {medplus.CLASSE_TAXA}

CONTA_CORRENTE = 'Omie.CASH'

# Contas a RECEBER: o Cliente somos nós (a associação). Como TODAS as filiais têm a
# mesma razão social, o que distingue o recebedor na OMIE é o CNPJ — então é o CNPJ
# que vai na coluna "Cliente" (chave de busca da OMIE). A filial também vai no
# Departamento, por legibilidade. (Diretoria 2026-06-23.)
CLIENTE_RECEBER = 'ASSOCIACAO SEUMED HOSPITAL DE OLHOS'

# De-para clínica (nome no MedPlus) -> CNPJ da filial SeuMed na OMIE.
# Obs.: o CNPJ .../0007-25 cobre PR2 (logradouro nº 819) e PR3 (nº 805) — mesmo CNPJ;
# nossa clínica "Maringá - Filial PR2 e PR3" usa esse. (Diretoria 2026-06-23.)
FILIAL_CNPJ = {
    'maringa - matriz':            '27.717.567/0001-30',
    'mandaguacu':                  '27.717.567/0002-10',
    'paicandu':                    '27.717.567/0003-00',
    'sarandi':                     '27.717.567/0004-82',
    'maringa - av brasil':         '27.717.567/0005-63',
    'mandaguari':                  '27.717.567/0006-44',
    'maringa - filial pr2 e pr3':  '27.717.567/0007-25',
}

# Departamento na OMIE (mesmo texto no a pagar e no a receber). O texto é "NN. Nome"
# (sem acento). PR2 e PR3 compartilham o código 07 (mesmo CNPJ .../0007-25), mas o
# Departamento distingue: a agenda de GLAUCOMA (ou nome com "PR3") é "07. PR3"; o resto
# da clínica "PR2 e PR3" é "07. PR2". (Diretoria 2026-06-25.)
DEPARTAMENTO = {
    'maringa - matriz':            '01. Matriz',
    'mandaguacu':                  '02. Mandaguacu',
    'paicandu':                    '03. Paicandu',
    'sarandi':                     '04. Sarandi',
    'maringa - av brasil':         '05. Brasil',
    'mandaguari':                  '06. Mandaguari',
}
_CLINICA_PR2_PR3 = 'maringa - filial pr2 e pr3'

# Categoria do a receber por grupo de convênio (diretoria: SUS / CISAMUSEP / PARTICULARES).
GRUPO_RECEBER = {
    'sus': 'SUS',
    'oci': 'SUS',                 # "OCI - SUS" é programa do SUS
    'cisa': 'CISAMUSEP',
    'particular': 'PARTICULARES',
    'convenio': 'PARTICULARES',   # parcerias/convênios privados entram como particulares
}
GRUPO_RECEBER_FALLBACK = 'PARTICULARES'


def _norm_clinica(nome: str) -> str:
    s = unicodedata.normalize('NFKD', str(nome or '')).encode('ascii', 'ignore').decode()
    return ' '.join(s.lower().split())


def cnpj_filial(clinica: str) -> str | None:
    """CNPJ da filial para a clínica (None se não houver de-para)."""
    return FILIAL_CNPJ.get(_norm_clinica(clinica))


def grupo_receber(convenio: str) -> str:
    """Grupo de categoria do a receber (SUS / CISAMUSEP / PARTICULARES)."""
    return GRUPO_RECEBER.get(regras.mapear_convenio(convenio), GRUPO_RECEBER_FALLBACK)


def clinica_pr2_pr3(clinica: str) -> bool:
    """A clínica é a 'Maringá - Filial PR2 e PR3' (que se desdobra em PR2/PR3)?"""
    return _norm_clinica(clinica) == _CLINICA_PR2_PR3


def departamento(clinica: str, subunidade: str = '') -> str | None:
    """Texto do Departamento OMIE ("01. Matriz", "07. PR3"...) para a clínica.

    Na clínica "PR2 e PR3", `subunidade` ('PR3'/'PR2') decide entre "07. PR3" e
    "07. PR2" (glaucoma/PR3 vai p/ PR3). None se a clínica não for mapeada."""
    if clinica_pr2_pr3(clinica):
        return '07. PR3' if (subunidade or '').upper() == 'PR3' else '07. PR2'
    return DEPARTAMENTO.get(_norm_clinica(clinica))

# Colunas (1-indexadas) no layout das planilhas OMIE; dados começam na linha 6.
LINHA_INICIAL = 6
COL_NOME = 3        # C: Fornecedor (pagar) / Cliente (receber)
COL_CATEGORIA = 4   # D
COL_CONTA = 5       # E
COL_VALOR = 6       # F
COL_REGISTRO = 10   # J: Data de Registro
COL_VENCIMENTO = 11 # K: Data de Vencimento
COL_OBSERVACOES = 19            # S: Observações (igual nos dois modelos)
COL_DEPARTAMENTO_PAGAR = 50     # Departamento (100%) — modelo a pagar
COL_DEPARTAMENTO_RECEBER = 42   # Departamento (100%) — modelo a receber

_RE_SUFIXO = re.compile(r'\s*\([^)]*\)\s*$')


def _sem_sufixo(nome: str) -> str:
    """Remove o sufixo de unidade entre parênteses do nome do profissional.

    "Dr. Carlos Eduardo (PR2)" -> "Dr. Carlos Eduardo" (a clínica já vai no
    Departamento; na observação fica só o nome do médico).
    """
    return _RE_SUFIXO.sub('', nome or '').strip()


@dataclass
class ResultadoSaida:
    nome_arquivo: str
    conteudo: bytes
    linhas: int
    pendencias: list = field(default_factory=list)


def venc_dia10_mes_seguinte(d: date) -> date:
    if d.month == 12:
        return date(d.year + 1, 1, 10)
    return date(d.year, d.month + 1, 10)


def _data_referencia(resultado) -> date:
    datas = [p.data for b in resultado.blocos for p in b.procedimentos if p.data]
    return max(datas) if datas else date.today()


def _fmt(d: date) -> str:
    return d.strftime('%d/%m/%Y')


def _escrever(modelo_path, linhas: list[dict], col_departamento: int) -> tuple[bytes, int]:
    """linhas: dicts com nome, categoria, valor, registro, vencimento, observacao, departamento.

    Preserva a formatação do template oficial: datas vão como VALOR de data real
    (não texto) com formato dd/mm/aaaa, e o valor como número 18,2 — para a OMIE
    importar sem erro de formato. A coluna de Departamento difere entre os dois
    modelos (a pagar e a receber), por isso vem como parâmetro.
    """
    wb = load_workbook(modelo_path)
    ws = wb[wb.sheetnames[0]]
    cols = (COL_NOME, COL_CATEGORIA, COL_CONTA, COL_VALOR, COL_REGISTRO,
            COL_VENCIMENTO, COL_OBSERVACOES, col_departamento)
    # Estilo da 1ª linha de dados do template — replicado em TODAS as linhas. Sem
    # isso, o modelo a pagar (que só traz ~3 linhas pré-formatadas) perde a
    # formatação a partir da 4ª linha; o a receber traz milhares e não sofria.
    estilo = {c: copy(ws.cell(row=LINHA_INICIAL, column=c)._style) for c in cols}
    r = LINHA_INICIAL
    for ln in linhas:
        for c in cols:
            ws.cell(row=r, column=c)._style = copy(estilo[c])
        ws.cell(row=r, column=COL_NOME, value=ln['nome'])
        ws.cell(row=r, column=COL_CATEGORIA, value=ln['categoria'])
        ws.cell(row=r, column=COL_CONTA, value=CONTA_CORRENTE)
        cval = ws.cell(row=r, column=COL_VALOR, value=round(float(ln['valor']), 2))
        cval.number_format = '0.00'
        for col, chave in ((COL_REGISTRO, 'registro'), (COL_VENCIMENTO, 'vencimento')):
            cel = ws.cell(row=r, column=col, value=ln[chave])  # objeto date, não string
            cel.number_format = 'DD/MM/YYYY'
        if ln.get('observacao'):
            ws.cell(row=r, column=COL_OBSERVACOES, value=ln['observacao'])
        if ln.get('departamento'):
            ws.cell(row=r, column=col_departamento, value=ln['departamento'])
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(linhas)


def gerar_contas_pagar(resultado, modelo_path) -> ResultadoSaida:
    """Uma linha por (médico × dia × clínica × classe).

    Cada bloco já representa um médico em um dia numa clínica; dentro do bloco
    somamos por classe (cirurgia/exame/preceptoria). Se o Dr. atende em duas
    clínicas em dois dias, saem quatro repasses — quatro linhas no a pagar.
    Observação = "Repasse dd/mm Nome do Dr."; Departamento = nome da clínica.
    """
    ref = _data_referencia(resultado)
    pendencias = []
    linhas = []
    for bloco in resultado.blocos:
        # Fornecedor (chave do a pagar na OMIE) = CNPJ do médico. (Diretoria 2026-06-28.)
        nome_forn = getattr(bloco, 'cnpj', '') or bloco.razao_social or bloco.profissional
        if not getattr(bloco, 'cnpj', ''):
            pendencias.append(f'{bloco.profissional}: sem CNPJ no cadastro (usei "{nome_forn}") — '
                              'confira na OMIE.')
        dia = bloco.data or ref
        venc = venc_dia10_mes_seguinte(dia)
        medico = _sem_sufixo(bloco.profissional)
        observacao = f'Repasse {dia.strftime("%d/%m")} {medico}'
        por_classe = defaultdict(float)    # classe -> soma honorário (precisão cheia)
        for p in bloco.procedimentos:
            if p.classe == medplus.CLASSE_TAXA:
                # Taxa de sala: por padrão é receita da clínica (só no a receber).
                # Entra no a pagar SÓ se o usuário informou na revisão um valor
                # devido ao médico — aí é repasse de cirurgia. (Diretoria 2026-06-23.)
                if p.status_calculo == 'calculado' and (p.honorario or 0) > 0:
                    por_classe[medplus.CLASSE_CIRURGIA] += p.honorario
                continue
            if p.status_calculo == 'calculado' and (p.honorario or 0) > 0:
                por_classe[p.classe] += p.honorario
            elif p.status_calculo == 'a_definir':
                pendencias.append(f'{bloco.profissional}: "{p.procedimento[:40]}" a definir — não entrou no a pagar.')
        for classe, soma in por_classe.items():
            categoria = CATEGORIA_POR_CLASSE.get(classe, '')
            if not categoria:
                # Sem categoria OMIE (ex.: "A classificar") NÃO entra no arquivo —
                # a OMIE rejeita linha sem categoria. Vira pendência para a pessoa
                # reclassificar na revisão antes de exportar.
                pendencias.append(f'{nome_forn}: R$ {soma:.2f} em "{classe}" — classifique '
                                  f'(Cirurgia/Exame/Preceptoria) na revisão; NÃO entrou no a pagar.')
                continue
            dep = departamento(bloco.clinica, getattr(bloco, 'subunidade', '')) or bloco.clinica
            linhas.append({'nome': nome_forn, 'categoria': categoria, 'valor': soma,
                           'registro': dia, 'vencimento': venc, 'observacao': observacao,
                           'departamento': dep})

    # Linhas dos anestesistas (categoria própria) — uma por atendimento/dia
    for a in getattr(resultado, 'anestesistas', []):
        dia = a.get('data') or ref
        anest = _sem_sufixo(a.get('anestesista', ''))
        cirurgiao = _sem_sufixo(a.get('cirurgiao', ''))
        # Observação traz o cirurgião-chefe entre parênteses (diretoria 2026-06-24):
        # "Repasse 22/06 Dra. Isabela Miwa Maeda (Dr. Rodolpho)".
        obs = f'Repasse {dia.strftime("%d/%m")} {anest}' + (f' ({cirurgiao})' if cirurgiao else '')
        dep = departamento(a.get('clinica', ''), a.get('subunidade', '')) or a.get('clinica', '')
        linhas.append({'nome': a.get('cnpj') or a.get('razao_social') or a.get('anestesista'),
                       'categoria': CATEGORIA_ANESTESISTA, 'valor': a.get('valor', 0),
                       'registro': dia, 'vencimento': venc_dia10_mes_seguinte(dia),
                       'observacao': obs, 'departamento': dep})

    linhas.sort(key=lambda x: (str(x['registro']), x['departamento'] or '', x['nome'], x['categoria']))
    conteudo, n = _escrever(modelo_path, linhas, COL_DEPARTAMENTO_PAGAR)
    return ResultadoSaida('OMIE_Contas_a_Pagar.xlsx', conteudo, n, pendencias)


def gerar_contas_pagar_de_linhas(linhas, modelo_path) -> ResultadoSaida:
    """Contas a pagar OMIE a partir de linhas JÁ PRONTAS (mesmo formato de _escrever:
    nome/categoria/valor/registro/vencimento/observacao/departamento). Usado para a
    preceptoria MENSAL, que não vem de um relatório da MedPlus."""
    ordenadas = sorted(linhas, key=lambda x: (str(x['registro']),
                                              x.get('departamento') or '', x['nome']))
    conteudo, n = _escrever(modelo_path, ordenadas, COL_DEPARTAMENTO_PAGAR)
    return ResultadoSaida('OMIE_Contas_a_Pagar.xlsx', conteudo, n, [])


# Nome amigável da classe para o RESUMO do relatório mensal.
RESUMO_CLASSE = {
    medplus.CLASSE_EXAME: 'Consultas e exames',
    medplus.CLASSE_CIRURGIA: 'Cirurgias e procedimentos',
    medplus.CLASSE_LAUDO: 'Laudos',
    medplus.CLASSE_PRECEPTORIA: 'Preceptoria',
}


def linhas_relatorio_pagar(resultado) -> list[dict]:
    """Linhas do a pagar em formato amigável (e JSON-ready) p/ o relatório mensal:
    {medico, clinica, classe, resumo, categoria, valor, vencimento, data} — uma por
    (médico × dia × clínica × classe) + os anestesistas. Datas em ISO (YYYY-MM-DD)."""
    ref = _data_referencia(resultado)
    out = []
    for bloco in resultado.blocos:
        dia = bloco.data or ref
        venc = venc_dia10_mes_seguinte(dia)
        medico = _sem_sufixo(bloco.profissional)
        clinica = bloco.clinica or ''
        dep = departamento(clinica, getattr(bloco, 'subunidade', '')) or clinica
        por_classe = defaultdict(float)
        for p in bloco.procedimentos:
            if p.classe == medplus.CLASSE_TAXA:
                if p.status_calculo == 'calculado' and (p.honorario or 0) > 0:
                    por_classe[medplus.CLASSE_CIRURGIA] += p.honorario
                continue
            if p.status_calculo == 'calculado' and (p.honorario or 0) > 0:
                por_classe[p.classe] += p.honorario
        for classe, soma in por_classe.items():
            out.append({'medico': medico, 'clinica': clinica, 'departamento': dep, 'classe': classe,
                        'resumo': RESUMO_CLASSE.get(classe, classe),
                        'categoria': CATEGORIA_POR_CLASSE.get(classe, classe),
                        'valor': round(soma, 2), 'vencimento': venc.isoformat(),
                        'data': dia.isoformat()})
    for a in getattr(resultado, 'anestesistas', []):
        dia = a.get('data') or ref
        clin = a.get('clinica', '') or ''
        out.append({'medico': _sem_sufixo(a.get('anestesista', '')), 'clinica': clin,
                    'departamento': departamento(clin, a.get('subunidade', '')) or clin,
                    'classe': 'Anestesia', 'resumo': 'Anestesia', 'categoria': CATEGORIA_ANESTESISTA,
                    'valor': round(float(a.get('valor') or 0), 2),
                    'vencimento': venc_dia10_mes_seguinte(dia).isoformat(), 'data': dia.isoformat()})
    return out


def gerar_contas_receber(resultado, modelo_path) -> ResultadoSaida:
    """Uma linha por (dia × clínica × grupo de convênio).

    Soma o valor bruto dos atendimentos daquele dia/clínica, separado por grupo
    (SUS / CISAMUSEP / PARTICULARES = categoria). Cliente = CNPJ da filial (chave de
    busca na OMIE, já que todas as filiais têm a mesma razão social); Departamento =
    clínica; Observação = "Recebimento dd/mm Clínica — GRUPO".
    """
    ref = _data_referencia(resultado)
    pendencias = []
    grupos = defaultdict(float)        # (dia, clínica, grupo) -> soma valor bruto
    sem_valor = 0
    sem_cnpj = set()
    conv_desconhecido = set()
    for bloco in resultado.blocos:
        clinica = bloco.clinica or ''
        # Departamento (com PR2/PR3 conforme a agenda) entra na chave — PR2 e PR3 saem
        # em linhas separadas mesmo na mesma clínica/dia/grupo.
        dep = departamento(clinica, getattr(bloco, 'subunidade', '')) or clinica
        # Repasse criado pelo sistema (agenda "Equipe Dr. Keiti"): não tem valor bruto
        # por natureza — não conta como pendência de "sem valor". (Diretoria 2026-06-28.)
        eh_sistema = getattr(bloco, 'equipe_keiti', False)
        for p in bloco.procedimentos:
            if p.classe == medplus.CLASSE_PRECEPTORIA:
                continue   # preceptoria é SÓ a pagar — não tem valor bruto nem avisa
            if p.valor is None:
                if not eh_sistema:
                    sem_valor += 1
                continue
            grupos[(bloco.data, clinica, dep, grupo_receber(p.convenio))] += p.valor
            # convênio que não casa em nenhum grupo (e não é taxa de sala) vai p/
            # PARTICULARES por padrão, mas avisa — igual ao "a definir" do a pagar.
            if p.classe != medplus.CLASSE_TAXA and regras.mapear_convenio(p.convenio) is None:
                conv_desconhecido.add((p.convenio or '').strip() or '(em branco)')
        if clinica and cnpj_filial(clinica) is None:
            sem_cnpj.add(clinica)
    if sem_valor:
        pendencias.append(f'{sem_valor} procedimento(s) sem valor bruto (arquivo do médico não traz o valor) — '
                          'use o relatório completo da MedPlus para o contas a receber.')
    for clinica in sorted(sem_cnpj):
        pendencias.append(f'Clínica "{clinica}" sem CNPJ de filial mapeado — Cliente saiu como a razão '
                          'social (a OMIE não consegue distinguir a filial). Cadastre o CNPJ.')
    for conv in sorted(conv_desconhecido):
        pendencias.append(f'Convênio "{conv}" não reconhecido — classificado como PARTICULARES no a '
                          'receber; confirme a categoria (SUS / CISAMUSEP / PARTICULARES).')

    linhas = []
    for (dia, clinica, dep, grupo), soma in sorted(grupos.items(),
                                                   key=lambda kv: (str(kv[0][0]), kv[0][2], kv[0][3])):
        d = dia or ref
        rotulo = f' {clinica}' if clinica else ''
        linhas.append({'nome': cnpj_filial(clinica) or CLIENTE_RECEBER, 'categoria': grupo,
                       'valor': soma, 'registro': d, 'vencimento': venc_dia10_mes_seguinte(d),
                       'observacao': f'Recebimento {d.strftime("%d/%m")}{rotulo} — {grupo}',
                       'departamento': dep})
    conteudo, n = _escrever(modelo_path, linhas, COL_DEPARTAMENTO_RECEBER)
    return ResultadoSaida('OMIE_Contas_a_Receber.xlsx', conteudo, n, pendencias)
