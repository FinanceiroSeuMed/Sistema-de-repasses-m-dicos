# -*- coding: utf-8 -*-
"""Relatório mensal compilado — junta os repasses (a pagar) do mês num único xlsx,
ordenado por médico, no formato "Repasses em Aberto" (anexo da diretoria):
colunas Filial | Destino | DataVencimento | Valor | Categoria, e um resumo por
classe (Consultas e exames / Cirurgias e procedimentos / Preceptoria / TOTAL).
"""

from __future__ import annotations

import io
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_ORDEM_RESUMO = ['Consultas e exames', 'Cirurgias e procedimentos', 'Laudos',
                 'Preceptoria', 'Anestesia', 'Ajuste']
# Rótulo exibido nos quadros de soma (capitalização pedida pela diretoria).
_LABEL_RESUMO = {
    'Consultas e exames': 'Consultas e Exames',
    'Cirurgias e procedimentos': 'Cirurgias e Procedimentos',
    'Laudos': 'Laudos',
    'Preceptoria': 'Preceptoria',
    'Anestesia': 'Anestesia',
    'Ajuste': 'Ajustes (+/−)',
}


def _chaves_resumo(por_classe):
    """Chaves na ordem padrão + qualquer classe fora dela (nenhum valor que entra no
    Total pode ficar invisível no quadro — auditoria 2026-07-02)."""
    extras = sorted(k for k in por_classe if k not in _ORDEM_RESUMO)
    return _ORDEM_RESUMO + extras
_MESES = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho',
          'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']


def nome_mes(aaaa_mm: str) -> str:
    """'2026-05' -> 'Maio 2026'."""
    try:
        ano, mes = aaaa_mm.split('-')
        return f'{_MESES[int(mes)]} {ano}'
    except Exception:
        return aaaa_mm


def _data(iso: str):
    try:
        return date.fromisoformat(iso)
    except Exception:
        return None


_HEADER = ['Filial', 'Destino', 'Data do Atendimento', 'Data de Vencimento', 'Valor', 'Categoria']
_MOEDA = '#,##0.00'
_DATA_FMT = 'DD/MM/YYYY'


def _quadro_somas(ws, linha0, titulo, por_classe, *, negrito, fill):
    """Escreve um quadro (rótulo | valor) nas colunas H/I a partir de `linha0`:
    cada classe com soma > 0 (na ordem padrão) + uma linha de Total. Devolve a
    última linha usada."""
    r = linha0
    cab = ws.cell(r, 8, titulo); cab.font = negrito; cab.fill = fill
    ws.cell(r, 9, '').fill = fill
    r += 1
    for k in _chaves_resumo(por_classe):
        if round(por_classe.get(k, 0), 2) != 0:
            ws.cell(r, 8, _LABEL_RESUMO.get(k, k))
            cel = ws.cell(r, 9, round(por_classe[k], 2)); cel.number_format = _MOEDA
            r += 1
    ws.cell(r, 8, 'Total').font = negrito
    cel = ws.cell(r, 9, round(sum(por_classe.values()), 2))
    cel.font = negrito; cel.number_format = _MOEDA
    return r


def gerar_relatorio_mensal(linhas: list[dict], titulo: str = 'Repasses em Aberto') -> bytes:
    """Gera o xlsx do mês. `linhas` = dicts de omie.linhas_relatorio_pagar.

    Layout (diretoria 2026-06-27): título; QUADRO GERAL (somas por classe + Total a
    Pagar) no topo; depois um grupo por médico — cada grupo repete o cabeçalho, lista
    seus lançamentos (com a Data do Atendimento) e traz AO LADO um quadro com as somas
    por classe (sem os zeros) + o total do médico; uma linha em branco separa os grupos.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contas a Pagar - Padrão'

    negrito = Font(bold=True)
    titulo_fnt = Font(bold=True, size=14)
    cab_fill = PatternFill('solid', fgColor='D9E1F2')      # cabeçalho das colunas
    geral_fill = PatternFill('solid', fgColor='FFE699')    # quadro geral (topo)
    medico_fill = PatternFill('solid', fgColor='E2EFDA')   # quadro por médico

    grupos = defaultdict(list)
    for ln in linhas:
        grupos[ln.get('medico') or '—'].append(ln)
    medicos = sorted(grupos, key=lambda m: m.lower())

    # Título
    ws.cell(1, 1, titulo).font = titulo_fnt

    # ---- Quadro GERAL (topo): somas por classe + Total a Pagar (todos os valores) ----
    geral = defaultdict(float)
    for ln in linhas:
        geral[ln.get('resumo') or 'Outros'] += float(ln.get('valor') or 0)
    r = 3
    cab = ws.cell(r, 1, 'RESUMO GERAL'); cab.font = negrito; cab.fill = geral_fill
    ws.cell(r, 2, '').fill = geral_fill
    r += 1
    for k in _chaves_resumo(geral):
        if round(geral.get(k, 0), 2) != 0:
            ws.cell(r, 1, _LABEL_RESUMO.get(k, k))
            cel = ws.cell(r, 2, round(geral[k], 2)); cel.number_format = _MOEDA
            r += 1
    ws.cell(r, 1, 'Total a Pagar').font = negrito
    cel = ws.cell(r, 2, round(sum(geral.values()), 2))
    cel.font = negrito; cel.number_format = _MOEDA
    r += 2   # linha em branco

    # ---- Um grupo por médico ----
    for medico in medicos:
        ls = sorted(grupos[medico], key=lambda x: x.get('data') or '')
        topo = r
        for c, t in enumerate(_HEADER, start=1):      # cabeçalho repetido no grupo
            cel = ws.cell(r, c, t); cel.font = negrito; cel.fill = cab_fill
        r += 1
        for ln in ls:
            ws.cell(r, 1, ln.get('departamento') or ln.get('clinica') or '')
            ws.cell(r, 2, ln.get('medico') or '')
            ca = ws.cell(r, 3, _data(ln.get('data'))); ca.number_format = _DATA_FMT
            cv = ws.cell(r, 4, _data(ln.get('vencimento'))); cv.number_format = _DATA_FMT
            cval = ws.cell(r, 5, round(float(ln.get('valor') or 0), 2)); cval.number_format = _MOEDA
            ws.cell(r, 6, ln.get('categoria') or '')
            r += 1
        # quadro de somas do médico, AO LADO (colunas H/I), a partir do cabeçalho do grupo
        por_classe = defaultdict(float)
        for ln in ls:
            por_classe[ln.get('resumo') or 'Outros'] += float(ln.get('valor') or 0)
        fim_quadro = _quadro_somas(ws, topo, medico, por_classe, negrito=negrito, fill=medico_fill)
        r = max(r, fim_quadro + 1) + 1   # garante espaço abaixo dos dados E do quadro

    for col, w in zip('ABCDEF', (18, 30, 18, 18, 13, 32)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions['H'].width = 26
    ws.column_dimensions['I'].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
