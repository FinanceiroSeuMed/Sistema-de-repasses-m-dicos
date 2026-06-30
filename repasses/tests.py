# -*- coding: utf-8 -*-
"""Suíte de regressão do motor de repasses.

Trava o comportamento validado (golds, casamento, classificação, OMIE) ANTES de
qualquer refatoração/otimização. Os testes de cálculo são SimpleTestCase (sem
banco): montam o livro a partir da planilha de regras e processam amostras reais.

Rodar:  python manage.py test repasses
"""

import io
import os
import unittest

from django.conf import settings
from django.test import Client, SimpleTestCase, TestCase

from repasses.services import medplus, omie, regras, repasse

_AMOSTRAS = r'C:\RepassesmedicosOMIE\amostras'
F18 = os.path.join(_AMOSTRAS, 'medplus_18_06.xls')             # Heric, com Valor (gold 4653.11)
FEXPORT = os.path.join(_AMOSTRAS, 'medplus_export.xls')        # 3 médicos (Tharcila gold 3250)
FFILIAIS = os.path.join(_AMOSTRAS, 'medplus_maio_filiais.xls')  # com coluna Clínica
FEXPORT2 = os.path.join(_AMOSTRAS, 'medplus_export_2.xls')      # tem OCI de residentes (Vida/Arthur)
F22 = os.path.join(_AMOSTRAS, 'medplus_22_06.xls')             # Heric Qtd/tono (gold Matriz 2449.11, PR2/PR3 1220)
PLANILHA = getattr(settings, 'REGRAS_REPASSE_PATH', '')


def _arquivos_presentes(*paths):
    return all(p and os.path.exists(p) for p in paths)


def _total_medico(rel, frag, clinica=None):
    """Soma dos honorários calculados (>0) do médico cujo nome contém frag.
    Se `clinica` for dado, soma só os procedimentos daquela clínica."""
    total = 0.0
    for b in rel.blocos:
        if frag in b.profissional.lower():
            for p in b.procedimentos:
                if clinica is not None and p.clinica != clinica:
                    continue
                if p.status_calculo == 'calculado' and (p.honorario or 0) > 0:
                    total += p.honorario
    return round(total, 2)


@unittest.skipUnless(_arquivos_presentes(PLANILHA), 'planilha de regras ausente')
class MotorRegrasTests(SimpleTestCase):
    """Casamento de convênio/procedimento/nome e classificação de valor."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.livro = regras.carregar_regras(PLANILHA)

    def test_mapear_convenio(self):
        casos = {
            'OCI - SUS': 'oci', 'Cisamusep': 'cisa', 'Sus Maringá': 'sus',
            'PG Glaucoma': 'sus', 'Bradesco': 'particular', 'Amil': 'particular',
            'Particular': 'particular', 'Parcerias I': 'particular',
        }
        for conv, esperado in casos.items():
            self.assertEqual(regras.mapear_convenio(conv), esperado, conv)

    def test_classificar_valor_vazio_e_traco(self):
        self.assertEqual(regras._classificar_valor(''), (regras.SEM_VALOR, None))
        self.assertEqual(regras._classificar_valor('-'), (regras.NEGATIVO, 0.0))

    def test_valor_db_roundtrip(self):
        self.assertAlmostEqual(regras._valor_db('28.5%'), 0.285)   # ponto decimal em %
        self.assertAlmostEqual(regras._valor_db('28,5%'), 0.285)
        self.assertAlmostEqual(regras._valor_db('24%'), 0.24)
        self.assertEqual(regras._valor_db('1.120'), 1120.0)        # ponto milhar em fixo
        self.assertAlmostEqual(regras._valor_db('549,99'), 549.99)
        self.assertEqual(regras._valor_db('-'), '-')
        self.assertIsNone(regras._valor_db(''))

    def test_medico_por_nome(self):
        flavia = self.livro.medico_por_nome('Flávia Zandonadi')   # abreviação casa
        self.assertIsNotNone(flavia)
        self.assertIn('flavia', regras.normalizar(flavia.nome))
        mm = self.livro.medico_por_nome('Maria Marta Cabral')     # não casa nome alheio
        self.assertFalse(mm and 'marilia' in regras.normalizar(mm.nome))
        th = self.livro.medico_por_nome('Dra. Tharcila Breginski da Rocha (PR2)')  # sufixo
        self.assertIsNotNone(th)

    def test_calcular_casos_conhecidos(self):
        def calc(p, c, v=None, m='Dr. Heric Sakamoto', q=1):
            return regras.calcular(self.livro, p, c, v, m, quantidade=q)
        self.assertEqual(calc('Consulta', 'Sus Maringá').honorario, 25.0)
        # Consulta SUS: diferencia tonografia (paciente pagou < R$30 -> R$10) da
        # consulta de avaliação normal (>= R$30 ou sem valor -> R$25). O texto
        # "(Tono)" no nome NÃO decide — só o valor pago. (Diretoria 2026-06-23.)
        self.assertEqual(calc('Consulta - Avaliação (Tono(1))', 'Sus Maringá', 12.0).honorario, 10.0)
        self.assertEqual(calc('Consulta - Avaliação (Tono(1))', 'Sus Maringá', 38.37).honorario, 25.0)
        # Qtd multiplica o valor fixo (mapeamento binocular Qtd=2 -> dobra).
        self.assertEqual(calc('Mapeamento de Retina', 'Sus Maringá', 98.48, q=2).honorario, 50.0)
        self.assertEqual(calc('Mapeamento de Retina', 'Sus Maringá', 98.48).honorario, 25.0)
        self.assertEqual(calc('Vitrectomia Posterior', 'Sus Maringá').honorario, 1000.0)
        self.assertEqual(calc('Vitrectomia Posterior com Infusão de Perfluocarbono',
                              'Sus Maringá').honorario, 1120.0)
        self.assertEqual(calc('Laudo - Retino/Angio', 'Sus Maringá').honorario, 10.0)
        self.assertEqual(calc('Laudo - Angio/Campi/Retino', 'Cisamusep').honorario, 30.0)
        self.assertAlmostEqual(calc('Sutura de Conjuntiva', 'Particular', 100.0).honorario, 24.0)
        # Copel e Sanepar seguem a regra de particular (consulta = R$60, como o particular)
        part = calc('Consulta em Oftalmologia', 'Particular', 150.0).honorario
        self.assertEqual(calc('Consulta em Oftalmologia', 'Copel', 150.0).honorario, part)
        self.assertEqual(calc('Consulta em Oftalmologia', 'Sanepar', 150.0).honorario, part)
        self.assertEqual(regras.mapear_convenio('Copel'), 'particular')
        self.assertEqual(regras.mapear_convenio('Sanepar'), 'particular')

    def test_faco_consulta_cisa_usa_base_cisa(self):
        # decisão da diretoria: CISA = base CISA (150) + consulta (25) = 175
        proc = ('FACOEMULSIFICACAO C/ IMPLANTE DE LENTE INTRA-OCULAR DOBRAVEL '
                '(Incluso: Pre-Operatorio: Consulta de Avaliacao)')
        r = regras.calcular(self.livro, proc, 'Cisamusep', None, 'Dr. Heric Sakamoto')
        self.assertEqual(r.honorario, 175.0)


@unittest.skipUnless(_arquivos_presentes(PLANILHA, F18, FEXPORT), 'amostras ausentes')
class GoldsTests(SimpleTestCase):
    """Golds de ponta-a-ponta (parser + motor), via planilha (sem banco)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.livro = regras.carregar_regras(PLANILHA)

    def _processado(self, arquivo):
        rel = medplus.ler_relatorio(arquivo)
        regras.processar(rel, self.livro)
        return rel

    def test_gold_heric_18_06(self):
        self.assertEqual(_total_medico(self._processado(F18), 'heric'), 4653.11)

    def test_gold_tharcila_export(self):
        self.assertEqual(_total_medico(self._processado(FEXPORT), 'tharcila'), 3250.0)

    @unittest.skipUnless(_arquivos_presentes(F22), 'amostra 22/06 ausente')
    def test_gold_heric_22_06(self):
        # Gold 22/06: Qtd (mapeamento binocular=2 -> R$50) e consulta-tono por valor
        # (avaliação SUS R$38,37 -> R$25, não R$10). Dois blocos por clínica.
        rel = self._processado(F22)
        self.assertEqual(_total_medico(rel, 'heric', 'Maringá - Matriz'), 2449.11)
        self.assertEqual(_total_medico(rel, 'heric', 'Maringá - Filial PR2 e PR3'), 1220.0)


@unittest.skipUnless(_arquivos_presentes(PLANILHA, F18), 'amostras ausentes')
class RepasseDocTests(SimpleTestCase):
    """Total e reconciliação (arredondamento) do documento de repasse."""

    def setUp(self):
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(F18)
        regras.processar(rel, livro)
        self.bloco = next(b for b in rel.blocos if 'heric' in b.profissional.lower())

    def test_linhas_somam_o_total(self):
        # Sem linha de "Arredondamento": a soma das linhas EXIBIDAS fecha com o Total
        # (a diferença de centavos foi embutida numa linha). Documento de conferência.
        cab, dados, total, n = repasse._linhas_repasse(self.bloco)
        self.assertEqual(total, 4653.11)
        soma_exibida = round(sum(linha[5] for linha in dados), 2)   # coluna Honorário
        self.assertEqual(soma_exibida, total)
        self.assertNotIn('Arredondamento', cab)


@unittest.skipUnless(_arquivos_presentes(PLANILHA, F22), 'amostra 22/06 ausente')
class RepasseLayoutTests(SimpleTestCase):
    """Excel e PDF do médico: mesmas colunas (com Paciente), célula de Total e SEM
    aviso de preceptoria nas exportações."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from openpyxl import load_workbook
        from pypdf import PdfReader
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(F22)
        regras.processar(rel, livro)
        cls.bloco = next(b for b in rel.blocos if 'heric' in b.profissional.lower())
        cls.bloco.lembrete = 'Repasse de preceptoria a lançar à parte: teste'  # não pode vazar p/ export
        cls.wb = load_workbook(io.BytesIO(repasse.gerar_excel(cls.bloco, 'Maringá')))
        cls.ws = cls.wb.active
        cls.pdf_txt = '\n'.join((p.extract_text() or '')
                                for p in PdfReader(io.BytesIO(repasse.gerar_pdf(cls.bloco, 'Maringá'))).pages)

    def _linha_cab(self):
        return next(r for r in range(1, 15) if self.ws.cell(r, 1).value == 'Data')

    def test_excel_colunas_e_total(self):
        cr = self._linha_cab()
        self.assertEqual([self.ws.cell(cr, c).value for c in range(1, 7)],
                         ['Data', 'Paciente', 'Procedimento', 'Convênio', 'Qtd.', 'Honorário'])
        # Total à DIREITA: rótulo "Total" na col 5 (Qtd) e valor na col 6 (Honorário).
        tot = next(r for r in range(cr, self.ws.max_row + 1) if self.ws.cell(r, 5).value == 'Total')
        esperado = repasse._total(self.bloco)
        self.assertAlmostEqual(float(self.ws.cell(tot, 6).value), esperado)
        pac = next(p.paciente for p in repasse.pagaveis(self.bloco) if p.paciente)
        self.assertEqual(self.ws.cell(cr + 1, 2).value, pac)   # 1ª linha traz o paciente

    def test_pdf_igual_ao_excel(self):
        self.assertIn('Paciente', self.pdf_txt)        # mesma coluna do Excel
        self.assertIn('Total', self.pdf_txt)
        self.assertNotIn('preceptoria a lançar', self.pdf_txt.lower())  # aviso só na preview


@unittest.skipUnless(_arquivos_presentes(PLANILHA, FFILIAIS), 'amostra de filiais ausente')
class OmiePagarTests(SimpleTestCase):
    """Estrutura do contas a pagar (sem linha de categoria vazia)."""

    def setUp(self):
        from openpyxl import load_workbook
        from repasses import views
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(FFILIAIS)
        views._filtrar_blocos(rel)
        views._separar_por_dia(rel)
        regras.processar(rel, livro)
        views._marcar_taxas_sala(rel)
        views._limpar_linhas(rel)
        rel.anestesistas = []
        conteudo = omie.gerar_contas_pagar(rel, settings.OMIE_PAGAR_TEMPLATE).conteudo
        wb = load_workbook(io.BytesIO(conteudo))
        self.ws = wb[wb.sheetnames[0]]

    def test_sem_categoria_vazia(self):
        r, vazias, linhas = omie.LINHA_INICIAL, 0, 0
        while self.ws.cell(r, omie.COL_NOME).value:
            linhas += 1
            if not self.ws.cell(r, omie.COL_CATEGORIA).value:
                vazias += 1
            r += 1
        self.assertGreater(linhas, 0)
        self.assertEqual(vazias, 0, 'nenhuma linha do a pagar pode sair sem categoria')


class OmieReceberTests(SimpleTestCase):
    """De-para clínica->CNPJ e categoria por grupo de convênio (a receber)."""

    def test_cnpj_e_grupo(self):
        self.assertEqual(omie.cnpj_filial('Maringá - Matriz'), '27.717.567/0001-30')
        self.assertEqual(omie.cnpj_filial('Mandaguaçu'), '27.717.567/0002-10')
        self.assertEqual(omie.cnpj_filial('Maringá - Filial PR2 e PR3'), '27.717.567/0007-25')
        self.assertIsNone(omie.cnpj_filial('Clínica Inexistente'))
        self.assertEqual(omie.grupo_receber('Sus Maringá'), 'SUS')
        self.assertEqual(omie.grupo_receber('OCI - SUS'), 'SUS')
        self.assertEqual(omie.grupo_receber('Cisamusep'), 'CISAMUSEP')
        self.assertEqual(omie.grupo_receber('Particular'), 'PARTICULARES')
        self.assertEqual(omie.grupo_receber('Parcerias I'), 'PARTICULARES')

    def test_departamento_texto(self):
        # Texto "NN. Nome" (sem acento), mesmo no a pagar e no a receber.
        self.assertEqual(omie.departamento('Maringá - Matriz'), '01. Matriz')
        self.assertEqual(omie.departamento('Mandaguaçu'), '02. Mandaguacu')
        self.assertEqual(omie.departamento('Paiçandu'), '03. Paicandu')
        self.assertEqual(omie.departamento('Sarandi'), '04. Sarandi')
        self.assertEqual(omie.departamento('Maringá - Av Brasil'), '05. Brasil')
        self.assertEqual(omie.departamento('Mandaguari'), '06. Mandaguari')
        # PR2 e PR3 (mesmo CNPJ 07): glaucoma/PR3 -> "07. PR3"; senão "07. PR2".
        self.assertEqual(omie.departamento('Maringá - Filial PR2 e PR3', 'PR3'), '07. PR3')
        self.assertEqual(omie.departamento('Maringá - Filial PR2 e PR3', 'PR2'), '07. PR2')
        self.assertEqual(omie.departamento('Maringá - Filial PR2 e PR3'), '07. PR2')
        self.assertIsNone(omie.departamento('Outra'))

    def test_deteccao_pr3(self):
        from repasses import views
        self.assertTrue(views._eh_pr3('Dra. Fulana (PR3)'))
        self.assertTrue(views._eh_pr3('Agenda Glaucoma'))
        self.assertTrue(views._eh_pr3('Dr. Beltrano - Glaucoma'))
        self.assertFalse(views._eh_pr3('Dra. Fulana (PR2)'))
        self.assertFalse(views._eh_pr3('Dr. Heric Sakamoto'))

    @unittest.skipUnless(_arquivos_presentes(PLANILHA, F22), 'amostra 22/06 ausente')
    def test_receber_22_06_estrutura(self):
        from openpyxl import load_workbook
        from repasses import views
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(F22)
        views._filtrar_blocos(rel); views._separar_por_dia(rel)
        regras.processar(rel, livro)
        rel.blocos = [b for b in rel.blocos if not regras.eh_residente(livro, b.profissional)]
        views._marcar_taxas_sala(rel); views._limpar_linhas(rel)
        res = omie.gerar_contas_receber(rel, settings.OMIE_RECEBER_TEMPLATE)
        wb = load_workbook(io.BytesIO(res.conteudo)); ws = wb[wb.sheetnames[0]]
        r, linhas = omie.LINHA_INICIAL, 0
        while ws.cell(r, omie.COL_NOME).value:
            linhas += 1
            cli = ws.cell(r, omie.COL_NOME).value
            cat = ws.cell(r, omie.COL_CATEGORIA).value
            self.assertRegex(str(cli), r'^\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}$')   # Cliente = CNPJ
            self.assertIn(cat, ('SUS', 'CISAMUSEP', 'PARTICULARES'))
            self.assertEqual(ws.cell(r, omie.COL_CONTA).value, omie.CONTA_CORRENTE)
            r += 1
        self.assertGreater(linhas, 0)


class OmiePagarCnpjTests(SimpleTestCase):
    """Fornecedor do contas a pagar = CNPJ do médico; sem CNPJ usa razão + avisa."""

    def _resultado(self, cnpj):
        from datetime import date
        from types import SimpleNamespace
        p = medplus.Procedimento(
            data=date(2026, 4, 14), data_texto='14/04/2026', paciente='Fulano',
            procedimento='Consulta', convenio='Particular', quantidade=1,
            valor=None, honorario_medplus=None, classe=medplus.CLASSE_EXAME)
        p.status_calculo = 'calculado'; p.honorario = 100.0
        b = medplus.BlocoMedico(profissional='Dr. Heric Sakamoto')
        b.procedimentos = [p]; b.data = date(2026, 4, 14); b.clinica = 'Maringá - Matriz'
        b.razao_social = 'HERIC MASSAAKI SAKAMOTO'; b.cnpj = cnpj
        return SimpleNamespace(blocos=[b], anestesistas=[])

    def test_fornecedor_e_cnpj(self):
        from openpyxl import load_workbook
        res = omie.gerar_contas_pagar(self._resultado('18.145.426/0001-07'),
                                      settings.OMIE_PAGAR_TEMPLATE)
        wb = load_workbook(io.BytesIO(res.conteudo)); ws = wb[wb.sheetnames[0]]
        self.assertEqual(ws.cell(omie.LINHA_INICIAL, omie.COL_NOME).value, '18.145.426/0001-07')
        self.assertFalse(any('sem CNPJ' in p for p in res.pendencias))

    def test_sem_cnpj_usa_razao_e_avisa(self):
        from openpyxl import load_workbook
        res = omie.gerar_contas_pagar(self._resultado(''), settings.OMIE_PAGAR_TEMPLATE)
        wb = load_workbook(io.BytesIO(res.conteudo)); ws = wb[wb.sheetnames[0]]
        self.assertEqual(ws.cell(omie.LINHA_INICIAL, omie.COL_NOME).value, 'HERIC MASSAAKI SAKAMOTO')
        self.assertTrue(any('sem CNPJ' in p for p in res.pendencias))


class OmieReceberKeitiTests(SimpleTestCase):
    """Repasse criado pelo sistema (Equipe Dr. Keiti) não avisa 'sem valor bruto'."""

    def _rel(self, equipe_keiti):
        from datetime import date
        from types import SimpleNamespace
        p = medplus.Procedimento(
            data=date(2026, 4, 14), data_texto='14/04/2026', paciente='Fulano',
            procedimento='Consulta', convenio='Particular', quantidade=1,
            valor=None, honorario_medplus=None, classe=medplus.CLASSE_EXAME)
        b = medplus.BlocoMedico(profissional='Dr. Keiti Fernando Shirasu')
        b.procedimentos = [p]; b.data = date(2026, 4, 14); b.clinica = 'Maringá - Matriz'
        b.equipe_keiti = equipe_keiti
        return SimpleNamespace(blocos=[b], anestesistas=[])

    def test_keiti_nao_avisa(self):
        res = omie.gerar_contas_receber(self._rel(True), settings.OMIE_RECEBER_TEMPLATE)
        self.assertFalse(any('sem valor bruto' in p for p in res.pendencias))

    def test_medico_comum_avisa(self):
        res = omie.gerar_contas_receber(self._rel(False), settings.OMIE_RECEBER_TEMPLATE)
        self.assertTrue(any('sem valor bruto' in p for p in res.pendencias))


class DataCurtaTests(SimpleTestCase):
    """Data abreviada DD/MM/AA para a tabela de revisão (cabe em uma linha)."""

    def _proc(self, data, texto):
        return medplus.Procedimento(
            data=data, data_texto=texto, paciente='', procedimento='', convenio='',
            quantidade=1, valor=None, honorario_medplus=None)

    def test_abrevia_pelo_date(self):
        from datetime import date
        self.assertEqual(self._proc(date(2026, 4, 14), '14/04/2026').data_curta, '14/04/26')

    def test_fallback_encurta_texto(self):
        self.assertEqual(self._proc(None, '14/04/2026').data_curta, '14/04/26')
        self.assertEqual(self._proc(None, '').data_curta, '')


class CategoriasOmieTests(SimpleTestCase):
    """Categorias OMIE do a pagar (diretoria 2026-06-30) + classe Laudos."""

    def test_categorias_por_classe(self):
        self.assertEqual(omie.CATEGORIA_POR_CLASSE[medplus.CLASSE_EXAME],
                         'Repasse Oftalmologistas - Consulta')
        self.assertEqual(omie.CATEGORIA_POR_CLASSE[medplus.CLASSE_CIRURGIA],
                         'Repasse Oftalmologistas - Cirurgia')
        self.assertEqual(omie.CATEGORIA_POR_CLASSE[medplus.CLASSE_LAUDO], 'Repasse Laudos')
        self.assertEqual(omie.CATEGORIA_POR_CLASSE[medplus.CLASSE_PRECEPTORIA], 'Preceptoria')
        self.assertEqual(omie.CATEGORIA_ANESTESISTA, 'Repasse Anestesiologistas')

    def test_laudo_palpite_e_subclasse(self):
        # "Laudo de OCT" casa como Laudo (não como Exame, mesmo contendo 'oct').
        self.assertEqual(medplus.classificar('Laudo de OCT'), medplus.CLASSE_LAUDO)
        self.assertEqual(medplus.classificar('Consulta em Oftalmologia'), medplus.CLASSE_EXAME)
        p = medplus.Procedimento(None, '', '', 'Laudo X', '', 1, None, None,
                                 classe=medplus.CLASSE_LAUDO)
        self.assertEqual(medplus.subclasse_preview(p), medplus.SUBCLASSE_LAUDO)

    def test_laudo_vira_repasse_laudos_no_a_pagar(self):
        from datetime import date
        from types import SimpleNamespace
        from openpyxl import load_workbook
        p = medplus.Procedimento(
            data=date(2026, 6, 27), data_texto='27/06/2026', paciente='F',
            procedimento='Laudo de Campimetria', convenio='Particular', quantidade=1,
            valor=None, honorario_medplus=None, classe=medplus.CLASSE_LAUDO)
        p.status_calculo = 'calculado'; p.honorario = 50.0
        b = medplus.BlocoMedico(profissional='Dr. Heric'); b.procedimentos = [p]
        b.data = date(2026, 6, 27); b.clinica = 'Maringá - Matriz'
        b.razao_social = ''; b.cnpj = '11.111.111/0001-11'
        rel = SimpleNamespace(blocos=[b], anestesistas=[])
        res = omie.gerar_contas_pagar(rel, settings.OMIE_PAGAR_TEMPLATE)
        wb = load_workbook(io.BytesIO(res.conteudo)); ws = wb[wb.sheetnames[0]]
        self.assertEqual(ws.cell(omie.LINHA_INICIAL, omie.COL_CATEGORIA).value, 'Repasse Laudos')


class PeriodoRotuloTests(SimpleTestCase):
    """Rótulos de período p/ nomes dos arquivos OMIE e do zip de repasses."""

    def test_rotulos(self):
        from datetime import date
        from repasses import views
        ini, fim = date(2026, 6, 27), date(2026, 6, 29)
        self.assertEqual(views._rotulo_periodo_dias(ini, fim), '27-29')
        self.assertEqual(views._rotulo_periodo_extenso(ini, fim), '27-29 de Junho')
        self.assertEqual(views._nome_omie('OMIE_Contas_a_Pagar', ini, fim),
                         'OMIE_Contas_a_Pagar_27-29 de Junho.xlsx')
        # um único dia
        self.assertEqual(views._rotulo_periodo_dias(ini, ini), '27')
        self.assertEqual(views._rotulo_periodo_extenso(ini, ini), '27 de Junho')
        # meses diferentes
        a, b = date(2026, 6, 30), date(2026, 7, 2)
        self.assertEqual(views._rotulo_periodo_dias(a, b), '30-06_a_02-07')
        self.assertEqual(views._rotulo_periodo_extenso(a, b), '30 de Junho a 02 de Julho')
        # sem datas -> nome OMIE sem sufixo
        self.assertEqual(views._nome_omie('OMIE_Contas_a_Receber', None, None),
                         'OMIE_Contas_a_Receber.xlsx')


@unittest.skipUnless(_arquivos_presentes(PLANILHA, F22), 'amostra 22/06 ausente')
class RelatorioMensalTests(SimpleTestCase):
    """Relatório mensal compilado por Dr. (formato Repasses em Aberto)."""

    def test_formato_e_resumo(self):
        from openpyxl import load_workbook
        from repasses.services import relatorio
        from repasses import views
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(F22)
        views._filtrar_blocos(rel); views._separar_por_dia(rel)
        regras.processar(rel, livro)
        rel.blocos = [b for b in rel.blocos if not regras.eh_residente(livro, b.profissional)]
        views._marcar_taxas_sala(rel); views._limpar_linhas(rel)

        linhas = omie.linhas_relatorio_pagar(rel)
        self.assertTrue(linhas)
        for ln in linhas:
            self.assertRegex(ln['data'], r'^\d{4}-\d{2}-\d{2}$')          # data ISO (p/ JSON)
            self.assertIn(ln['resumo'], ('Consultas e exames', 'Cirurgias e procedimentos',
                                         'Preceptoria', 'Anestesia'))
        conteudo = relatorio.gerar_relatorio_mensal(linhas, 'Teste Mensal')
        ws = load_workbook(io.BytesIO(conteudo))['Contas a Pagar - Padrão']
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertIn('Teste Mensal', textos)                 # título no topo
        self.assertIn('RESUMO GERAL', textos)                 # quadro geral (E5)
        self.assertIn('Data do Atendimento', textos)          # nova coluna (E2)
        self.assertIn('Data de Vencimento', textos)           # renomeada (E1)
        self.assertNotIn('DataVencimento', textos)            # nome antigo sumiu
        # cabeçalho repetido em cada grupo de médico (E3)
        n_medicos = len({l['medico'] for l in linhas})
        self.assertGreaterEqual(textos.count('Filial'), n_medicos)
        # Total a Pagar geral = soma de TODOS os valores (E5)
        total_esperado = round(sum(float(l['valor']) for l in linhas), 2)
        achou = False
        for row in ws.iter_rows():
            rotulos = [c.value for c in row]
            if 'Total a Pagar' in rotulos:
                i = rotulos.index('Total a Pagar')
                self.assertAlmostEqual(round(float(row[i + 1].value), 2), total_esperado)
                achou = True
        self.assertTrue(achou)


@unittest.skipUnless(_arquivos_presentes(PLANILHA, FEXPORT2), 'amostra com OCI de residente ausente')
class OciResidentesTests(SimpleTestCase):
    """OCI feito por residente vai para o repasse do Dr. Alessander."""

    def test_oci_residente_vai_para_alessander(self):
        from repasses import views
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(FEXPORT2)
        views._filtrar_blocos(rel)
        views._separar_por_dia(rel)
        regras.processar(rel, livro)
        raw_res = sum(1 for b in rel.blocos if regras.eh_residente(livro, b.profissional)
                      for p in b.procedimentos if regras.mapear_convenio(p.convenio) == 'oci')
        self.assertGreater(raw_res, 0, 'o teste precisa de OCI em agenda de residente')

        views._oci_residentes(rel, livro)
        rel.blocos = [b for b in rel.blocos if not regras.eh_residente(livro, b.profissional)]

        movidas = [(b, p) for b in rel.blocos for p in b.procedimentos
                   if 'residente' in (p.motivo_calculo or '').lower()]
        self.assertEqual(len(movidas), raw_res, 'todo OCI de residente deve ser transferido (nada perdido)')
        # todas no bloco do Alessander e com honorário calculado (> 0)
        for b, p in movidas:
            self.assertIn('alessander', regras.normalizar(b.profissional))
            self.assertGreater(p.honorario or 0, 0)


class ParserUnitTests(SimpleTestCase):
    """Funções puras do parser (sem arquivos)."""

    def test_para_numero(self):
        self.assertAlmostEqual(medplus._para_numero('1.020,60'), 1020.6)
        self.assertAlmostEqual(medplus._para_numero('1020.6'), 1020.6)
        self.assertAlmostEqual(medplus._para_numero('25'), 25.0)
        self.assertIsNone(medplus._para_numero(''))
        self.assertIsNone(medplus._para_numero(None))

    def test_eh_cirurgia(self):
        self.assertTrue(medplus.eh_cirurgia('Facoemulsificação + LIO'))
        self.assertTrue(medplus.eh_cirurgia('Vitrectomia Posterior'))
        self.assertFalse(medplus.eh_cirurgia('Capsulotomia a YAG Laser'))
        self.assertFalse(medplus.eh_cirurgia('Consulta em Oftalmologia'))

    def test_subclasse_preview(self):
        def p(nome, classe):
            return medplus.Procedimento(None, '', '', nome, '', 1, None, None, classe=classe)
        self.assertEqual(medplus.subclasse_preview(p('Facoemulsificação', medplus.CLASSE_CIRURGIA)),
                         medplus.SUBCLASSE_CIRURGIA)
        self.assertEqual(medplus.subclasse_preview(p('YAG Laser', medplus.CLASSE_CIRURGIA)),
                         medplus.SUBCLASSE_PROCEDIMENTO)
        self.assertEqual(medplus.subclasse_preview(p('Consulta', medplus.CLASSE_EXAME)),
                         medplus.SUBCLASSE_EXAME)
        self.assertEqual(medplus.subclasse_preview(p('Facectomia', medplus.CLASSE_TAXA)),
                         medplus.SUBCLASSE_TAXA)


@unittest.skipUnless(_arquivos_presentes(PLANILHA, F22), 'amostra 22/06 ausente')
class CorrecaoPorMedicoTests(TestCase):
    """Correção memorizada por médico: vale só p/ o médico salvo (e os marcados)."""

    def test_isolada_por_medico(self):
        from repasses.services import correcoes
        from repasses import views
        correcoes.memorizar('Consulta em Oftalmologia', 'Particular', 999.0,
                            medico='Dr. Heric Sakamoto')
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(F22)
        views._filtrar_blocos(rel); views._separar_por_dia(rel)
        regras.processar(rel, livro)
        correcoes.aplicar(rel)

        def consultas_particular(frag):
            return [p.honorario for b in rel.blocos if frag in b.profissional.lower()
                    for p in b.procedimentos
                    if p.procedimento.strip().lower() == 'consulta em oftalmologia'
                    and (p.convenio or '').strip().lower() == 'particular']
        her = consultas_particular('heric')
        rob = consultas_particular('roberta')
        self.assertTrue(her and all(abs(h - 999.0) < 0.01 for h in her))   # Heric: todas 999
        self.assertTrue(rob and all(abs(h - 999.0) > 0.01 for h in rob))   # Roberta: nenhuma 999

    def test_casa_por_cadastro_mesmo_nome_diferente(self):
        # Bug pego na revisão: o modal lista o nome do CADASTRO (ex.: "Dra. Tharcila"),
        # que difere do nome do MedPlus ("Dra. Tharcila Breginski da Rocha"). Sem
        # resolver pelo cadastro, a correção do "outro médico" nunca casaria.
        from repasses.services import correcoes
        from repasses import views
        livro = regras.carregar_regras(PLANILHA)
        rel = medplus.ler_relatorio(F22)
        views._filtrar_blocos(rel); views._separar_por_dia(rel)
        regras.processar(rel, livro)
        b = next(x for x in rel.blocos if 'tharcila' in x.profissional.lower())
        canon = livro.medico_por_nome(b.profissional).nome
        self.assertNotEqual(canon, b.profissional)         # cadastro != MedPlus
        proc = next(p.procedimento for p in b.procedimentos if 'facoemulsifica' in p.procedimento.lower())
        correcoes.memorizar(proc, 'Sus Maringá', 777.0, medico=canon)   # como o modal salvaria

        rel2 = medplus.ler_relatorio(F22)
        views._filtrar_blocos(rel2); views._separar_por_dia(rel2)
        regras.processar(rel2, livro)
        correcoes.aplicar(rel2, livro)                     # resolve o médico pelo cadastro
        thar = [p.honorario for x in rel2.blocos if 'tharcila' in x.profissional.lower()
                for p in x.procedimentos if 'facoemulsifica' in p.procedimento.lower()
                and (p.convenio or '').strip().lower().startswith('sus')]
        self.assertTrue(thar and all(abs(h - 777.0) < 0.01 for h in thar))


class FellowSplitTests(TestCase):
    """Catarata com fellow: 60/40 calculado e EDITÁVEL (override do chefe e do fellow)."""

    def _rel_catarata(self, valor=1000.0):
        from types import SimpleNamespace
        p = medplus.Procedimento(None, '', 'Paciente X', 'Facoemulsificação', 'Particular', 1, valor, None)
        p.idx = 5
        p.status_calculo = regras.CATARATA
        b = medplus.BlocoMedico(profissional='Dr. Chefe')
        b.procedimentos = [p]
        return SimpleNamespace(blocos=[b]), p

    def _fellow_line(self, rel):
        return next(x for bl in rel.blocos for x in bl.procedimentos if getattr(x, 'sintetica', False))

    def test_calculado_60_40(self):
        from repasses import views
        rel, p = self._rel_catarata(1000.0)   # à vista 30% -> total 300; chefe 180, fellow 120
        views._resolver_cirurgias(rel, {'cat_modo_5': 'avista', 'cat_fellow_5': 'Dr. Fellow'})
        self.assertAlmostEqual(p.honorario, 180.0)
        self.assertAlmostEqual(self._fellow_line(rel).honorario, 120.0)

    def test_chefe_override_no_proprio_campo(self):
        from repasses import views
        rel, p = self._rel_catarata(10000.0)   # à vista 30% -> total 3000
        views._resolver_cirurgias(rel, {'cat_modo_5': 'avista', 'cat_fellow_5': 'Dr. Fellow',
                                        'cat_chefe_5': '2000'})
        self.assertAlmostEqual(p.honorario, 2000.0)                       # chefe sobrescrito
        self.assertAlmostEqual(self._fellow_line(rel).honorario, 1200.0)  # fellow segue 40%

    def test_fellow_override_na_propria_linha(self):
        from repasses import views
        rel, p = self._rel_catarata(10000.0)   # total 3000; chefe 1800, fellow 1200
        post = {'cat_modo_5': 'avista', 'cat_fellow_5': 'Dr. Fellow'}
        views._resolver_cirurgias(rel, post)
        views._indexar_sinteticas(rel)            # dá idx à linha do fellow
        fl = self._fellow_line(rel)
        post[f'hon_{fl.idx}'] = '900'             # edita o fellow NA LINHA DELE
        views._aplicar_edicoes_sinteticas(rel, post)
        self.assertAlmostEqual(fl.honorario, 900.0)
        self.assertAlmostEqual(p.honorario, 1800.0)  # chefe segue 60% (independente)

    def test_chefe_negativo_ignorado(self):
        from repasses import views
        rel, p = self._rel_catarata(10000.0)   # -500 inválido -> usa 60%
        views._resolver_cirurgias(rel, {'cat_modo_5': 'avista', 'cat_fellow_5': 'Dr. Fellow',
                                        'cat_chefe_5': '-500'})
        self.assertAlmostEqual(p.honorario, 1800.0)

    def test_assistente_nao_pede_anestesista(self):
        # Com assistente, o bloco do assistente (linha sintética) NÃO ativa anestesista.
        from repasses import views
        rel, p = self._rel_catarata(10000.0)
        views._resolver_cirurgias(rel, {'cat_modo_5': 'avista', 'cat_fellow_5': 'Dr. Fellow'})
        bloco_assist = next(b for b in rel.blocos if regras.normalizar(b.profissional) == 'dr fellow')
        self.assertTrue(getattr(bloco_assist, 'participacao', False))
        self.assertFalse(bloco_assist.tem_cirurgia)   # não pergunta anestesista de novo

    def test_sufixo_cirurgiao_assistente(self):
        from repasses import views
        rel, p = self._rel_catarata(10000.0)
        views._resolver_cirurgias(rel, {'cat_modo_5': 'avista', 'cat_fellow_5': 'Dr. Fellow'})
        self.assertEqual(getattr(p, 'sufixo_export', ''), ' - Cirurgião 60%')
        self.assertEqual(getattr(self._fellow_line(rel), 'sufixo_export', ''), ' - Assistente 40%')

    def test_cirurgiao_sem_assistente_sem_sufixo(self):
        from repasses import views
        rel, p = self._rel_catarata(10000.0)
        views._resolver_cirurgias(rel, {'cat_modo_5': 'avista', 'cat_fellow_5': '__sem__'})
        self.assertEqual(getattr(p, 'sufixo_export', ''), '')   # sem split -> sem adendo
        self.assertAlmostEqual(p.honorario, 3000.0)             # cirurgião 100%


class LoteNomeFiliaisTests(TestCase):
    """Nome amigável ('Repasse médico - data') e resumo de filiais no histórico."""

    def _lote(self, linhas, tok='hist'):
        from repasses.models import Lote
        return Lote.objects.create(token=tok, linhas_pagar=linhas)

    def test_nome_repasse(self):
        l = self._lote([])
        self.assertTrue(l.nome_repasse.startswith('Repasse médico - '))

    def test_filiais_todas(self):
        from repasses.services.omie import FILIAL_CNPJ
        raw = ['Maringá - Matriz', 'Mandaguaçu', 'Paiçandu', 'Sarandi',
               'Maringá - Av Brasil', 'Mandaguari', 'Maringá - Filial PR2 e PR3']
        self.assertEqual(len(raw), len(FILIAL_CNPJ))
        linhas = [{'clinica': c, 'departamento': '01. X'} for c in raw]
        self.assertEqual(self._lote(linhas, 't1').filiais_resumo, 'Todas')

    def test_filiais_lista_por_codigo(self):
        linhas = [{'clinica': 'Maringá - Filial PR2 e PR3', 'departamento': '07. PR2'},
                  {'clinica': 'Maringá - Matriz', 'departamento': '01. Matriz'}]
        self.assertEqual(self._lote(linhas, 't2').filiais_resumo, 'Matriz, PR2')   # ordena por código


class VitrectomiaTests(TestCase):
    """Vitrectomia Anterior tratada EXATAMENTE como catarata (mesmos valores)."""

    def _regras(self):
        from repasses.models import RegraRepasse
        from repasses.management.commands.seed_regras import Command
        cls = 'Cirurgias e Procedimentos'
        RegraRepasse.objects.create(nome='Cirurgia de Catarata - Rodolpho', classe=cls,
                                    val_sus='160', val_cisa='160')
        RegraRepasse.objects.create(nome='Cirurgia de Catarata - Particular à vista', classe=cls,
                                    val_particular='30%')
        RegraRepasse.objects.create(nome='Vitrectomia Anterior', classe=cls,
                                    val_particular='mesmo valor da catarata')
        Command()._vitrectomia_como_catarata()

    def test_seed_espelha_catarata_e_remove_bogus(self):
        from repasses.models import RegraRepasse
        self._regras()
        self.assertFalse(RegraRepasse.objects.filter(nome_norm='vitrectomia anterior').exists())
        esp = RegraRepasse.objects.get(nome='Vitrectomia Anterior - Rodolpho')
        self.assertEqual((esp.val_sus, esp.val_cisa), ('160', '160'))

    def test_engine_trata_como_catarata(self):
        self._regras()
        livro = regras.carregar_livro_padrao()
        # particular -> precisa de à vista/parcelado + fellow (status catarata)
        self.assertEqual(regras.calcular(livro, 'Vitrectomia Anterior', 'Particular',
                                         5000.0, 'Dr. Rodolpho').status, regras.CATARATA)
        # SUS -> valor fixo do médico (mesmo da catarata)
        self.assertEqual(regras.calcular(livro, 'Vitrectomia Anterior', 'SUS',
                                         5000.0, 'Dr. Rodolpho').honorario, 160.0)


class EquipeDestinoTests(TestCase):
    """Agenda 'Equipe Dr. Keiti' roteada para o médico escolhido (Keiti/Thalia)."""

    def test_resolver_equipe_roteia(self):
        from types import SimpleNamespace
        from repasses import views
        from repasses.models import Medico
        Medico.objects.create(nome='Dr. Keiti Fernando Shirasu', razao_social='CLINICA SHIRASU')
        b = medplus.BlocoMedico(profissional=views._NOME_EQUIPE_KEITI); b.equipe_keiti = True
        rel = SimpleNamespace(blocos=[b])
        views._resolver_equipe(rel, {'equipe_destino_0': 'Dr. Keiti Fernando Shirasu'})
        self.assertEqual(b.profissional, 'Dr. Keiti Fernando Shirasu')
        self.assertEqual(b.razao_social, 'CLINICA SHIRASU')

    def test_sem_escolha_permanece_equipe(self):
        from types import SimpleNamespace
        from repasses import views
        from repasses.models import Medico
        Medico.objects.create(nome='Dr. Keiti Fernando Shirasu')
        b = medplus.BlocoMedico(profissional=views._NOME_EQUIPE_KEITI); b.equipe_keiti = True
        rel = SimpleNamespace(blocos=[b])
        views._resolver_equipe(rel, {'equipe_destino_0': '__verificar__'})
        self.assertEqual(b.profissional, views._NOME_EQUIPE_KEITI)   # pendente


class DuplicidadeTests(TestCase):
    """Bloqueio de lote duplicado (mesmo período já exportado)."""

    def _rel_com(self, *fps_de):
        from types import SimpleNamespace
        blocos = []
        for prof, proc, data in fps_de:
            p = medplus.Procedimento(data, '', 'Pac', proc, 'Particular', 1, 100, None)
            p.honorario = 100.0; p.status_calculo = 'calculado'
            b = medplus.BlocoMedico(profissional=prof); b.procedimentos = [p]
            blocos.append(b)
        return SimpleNamespace(blocos=blocos)

    def test_bloqueia_duplicado_e_libera_distinto(self):
        import datetime
        from repasses import views
        from repasses.models import Lote
        rel = self._rel_com(('Dr. A', 'Consulta', datetime.date(2026, 6, 16)),
                            ('Dr. B', 'Exame', datetime.date(2026, 6, 16)))
        fps = views._fingerprints_pagaveis(rel)
        Lote.objects.create(token='lote-antigo', arquivo_nome='junho.xls', fingerprints=fps)
        # outro upload com os MESMOS atendimentos -> duplicado
        self.assertTrue(views._lotes_duplicados('novo-token', fps))
        # re-exportar o MESMO upload (mesmo token) não bloqueia
        self.assertFalse(views._lotes_duplicados('lote-antigo', fps))
        # mesmo dia/mês de OUTRO ano -> NÃO casa (fingerprint tem o ano)
        rel2027 = self._rel_com(('Dr. A', 'Consulta', datetime.date(2027, 6, 16)),
                                ('Dr. B', 'Exame', datetime.date(2027, 6, 16)))
        self.assertFalse(views._lotes_duplicados('novo-token', views._fingerprints_pagaveis(rel2027)))


class SeedNomeTests(SimpleTestCase):
    def test_corrige_redientes(self):
        from repasses.management.commands.seed_medicos import _corrigir_nome
        self.assertEqual(_corrigir_nome('Redientes Dra. Regina'), 'Residentes Dra. Regina')
        self.assertEqual(_corrigir_nome('Dr. Rodolpho'), 'Dr. Rodolpho')   # inalterado


class ClasseMemorizadaTests(TestCase):
    """Classe memorizada por procedimento (global, todos os médicos)."""

    def _linha(self, proc, classe, sugerida):
        from types import SimpleNamespace
        p = medplus.Procedimento(None, '', 'Pac', proc, 'Particular', 1, 100, None)
        p.classe = classe; p.classe_sugerida = sugerida
        p.honorario = 20.0; p.status_calculo = 'calculado'
        b = medplus.BlocoMedico(profissional='Dr. X'); b.procedimentos = [p]
        return SimpleNamespace(blocos=[b], unidade=''), p

    def test_memoriza_quando_classifica(self):
        from repasses import views
        from repasses.models import ClasseMemorizada
        rel, p = self._linha('Retirada de Corpo Estranho', 'Exames e Consultas',
                             medplus.CLASSE_INDEFINIDA)
        self.assertTrue(views._memorizar_classes(rel, {}))
        c = ClasseMemorizada.objects.get(proc_norm=regras.normalizar('Retirada de Corpo Estranho'))
        self.assertEqual(c.classe, 'Exames e Consultas')

    def test_nao_memoriza_sem_mudanca(self):
        from repasses import views
        from repasses.models import ClasseMemorizada
        rel, p = self._linha('Consulta', 'Exames e Consultas', 'Exames e Consultas')
        views._memorizar_classes(rel, {})
        self.assertEqual(ClasseMemorizada.objects.count(), 0)

    def test_aplicar_reaplica_para_qualquer_medico(self):
        from types import SimpleNamespace
        from repasses.services import correcoes
        correcoes.memorizar_classe('Laudo Especial', 'Exames e Consultas')
        p = medplus.Procedimento(None, '', '', 'Laudo Especial', 'SUS', 1, 100, None)
        p.classe = medplus.CLASSE_INDEFINIDA; p.status_calculo = 'calculado'
        b = medplus.BlocoMedico(profissional='Dra. Outra'); b.procedimentos = [p]
        n = correcoes.aplicar_classes(SimpleNamespace(blocos=[b]))
        self.assertEqual(n, 1)
        self.assertEqual(p.classe, 'Exames e Consultas')   # vale p/ qualquer médico

    def test_nao_reativa_classe_desligada(self):
        # Re-exportar um lote antigo NÃO pode ressuscitar uma classe que foi desligada.
        from repasses import views
        from repasses.models import ClasseMemorizada
        ClasseMemorizada.objects.create(procedimento='Proc X', classe='Exames e Consultas',
                                        ativo=False)
        rel, p = self._linha('Proc X', 'Exames e Consultas', medplus.CLASSE_INDEFINIDA)
        views._memorizar_classes(rel, {})
        self.assertFalse(ClasseMemorizada.objects.get(procedimento='Proc X').ativo)

    def test_taxa_de_sala_nao_memoriza_nem_sobrescreve(self):
        from types import SimpleNamespace
        from repasses import views
        from repasses.models import ClasseMemorizada
        from repasses.services import correcoes
        # memória global remapeando o MESMO nome p/ cirurgia
        correcoes.memorizar_classe('Sala Especial', 'Cirurgias e Procedimentos')
        # linha de TAXA (convênio "Taxas De Sala") com o mesmo nome
        p = medplus.Procedimento(None, '', '', 'Sala Especial', 'Taxas De Sala', 1, 0, None)
        p.classe = medplus.CLASSE_TAXA; p.status_calculo = 'a_definir'
        b = medplus.BlocoMedico(profissional='Dr. Z'); b.procedimentos = [p]
        rel = SimpleNamespace(blocos=[b], unidade='')
        correcoes.aplicar_classes(rel)
        self.assertEqual(p.classe, medplus.CLASSE_TAXA)      # NÃO sobrescreveu a taxa
        p.classe_sugerida = medplus.CLASSE_TAXA; p.classe = 'Cirurgias e Procedimentos'
        antes = ClasseMemorizada.objects.count()
        views._memorizar_classes(rel, {})
        self.assertEqual(ClasseMemorizada.objects.count(), antes)  # taxa não vira memória


class ReginaTests(SimpleTestCase):
    """Dra. Regina: valor por faixa + distinção indivíduo x equipe."""

    def test_valor_por_faixa(self):
        from repasses import views
        self.assertEqual(views._valor_regina(1), 1000.0)
        self.assertEqual(views._valor_regina(23), 1000.0)
        self.assertEqual(views._valor_regina(24), 1500.0)   # a partir de 24
        self.assertEqual(views._valor_regina(40), 1500.0)

    def test_individuo_vs_equipe(self):
        from repasses import views
        self.assertTrue(views._eh_regina_dinheiro('Dra. Regina'))
        self.assertFalse(views._eh_regina_dinheiro('Equipe Dra. Regina'))
        self.assertFalse(views._eh_regina_dinheiro('Redientes Dra. Regina'))
        self.assertFalse(views._eh_regina_dinheiro('Dra. Isabela Miwa Maeda'))


class KeitiOciEquipeTests(SimpleTestCase):
    """Agenda 'Equipe Dr. Keiti' -> mantida + marcada (destino Keiti/Thalia na revisão)."""

    def test_reconhece_equipe_keiti(self):
        from repasses import views
        self.assertTrue(views._eh_equipe_keiti('equipe - dr. keiti shirasu'))
        self.assertFalse(views._eh_equipe_keiti('dra. thalia macaris'))
        self.assertFalse(views._eh_equipe_keiti('equipe dra. regina'))

    def test_filtrar_mantem_e_marca_equipe_keiti(self):
        from types import SimpleNamespace
        from repasses import views
        b = medplus.BlocoMedico(profissional='Equipe - Dr. Keiti Shirasu')
        rel = SimpleNamespace(blocos=[b])
        views._filtrar_blocos(rel)
        self.assertEqual(len(rel.blocos), 1)               # não é descartada
        self.assertTrue(getattr(rel.blocos[0], 'equipe_keiti', False))
        self.assertEqual(rel.blocos[0].profissional, views._NOME_EQUIPE_KEITI)

    def test_aplicar_keiti_pula_equipe(self):
        # _aplicar_keiti não mexe na agenda da equipe (destino é escolhido depois).
        from types import SimpleNamespace
        from repasses import views
        oci = medplus.Procedimento(None, '', '', 'OCI AVALIAÇÃO DE ESTRABISMO',
                                   'OCI - SUS', 1, 200, None)
        oci.classe = medplus.CLASSE_EXAME; oci.honorario = 50.0; oci.status_calculo = 'calculado'
        b = medplus.BlocoMedico(profissional=views._NOME_EQUIPE_KEITI); b.procedimentos = [oci]
        b.equipe_keiti = True
        views._aplicar_keiti(SimpleNamespace(blocos=[b]))
        self.assertEqual([p.honorario for p in b.procedimentos], [50.0])   # intacto

    def test_aplicar_keiti_oci_mantem_valor(self):
        from repasses import views
        oci = medplus.Procedimento(None, '', '', 'OCI AVALIAÇÃO DE ESTRABISMO',
                                   'OCI - SUS', 1, 200, None)
        oci.classe = medplus.CLASSE_EXAME           # OCI vem como "Exames e Consultas"
        oci.honorario = 50.0; oci.status_calculo = 'calculado'
        oci.motivo_calculo = 'Valor fixo (regra "Avaliação de Estrabismo").'
        exame = medplus.Procedimento(None, '', '', 'Consulta', 'SUS', 1, 0, None)
        exame.classe = medplus.CLASSE_EXAME
        b = medplus.BlocoMedico(profissional='Dr. Keiti Fernando Shirasu')
        b.procedimentos = [oci, exame]
        from types import SimpleNamespace
        views._aplicar_keiti(SimpleNamespace(blocos=[b]))
        valores = [p.honorario for p in b.procedimentos]
        self.assertIn(50.0, valores)        # OCI manteve o próprio valor
        self.assertIn(1000.0, valores)      # a consulta (não-OCI) virou pacote de R$1.000


class CamposObrigatoriosTests(SimpleTestCase):
    """Campos a verificar (classe/forma de pagamento/fellow/anestesista) bloqueiam o export."""

    def _rel(self):
        from types import SimpleNamespace
        p = medplus.Procedimento(None, '', 'Pac', 'Facoemulsificação', 'Particular', 1, 1000, None)
        p.idx = 1
        p.classe = medplus.CLASSE_CIRURGIA      # tem_cirurgia -> exige anestesista
        p.status_calculo = regras.CATARATA      # exige forma de pagamento + fellow
        b = medplus.BlocoMedico(profissional='Dr. X'); b.procedimentos = [p]
        return SimpleNamespace(blocos=[b])

    def test_pendente_quando_nao_confirmado(self):
        from repasses import views
        campos = dict(views._campos_pendentes(self._rel(), {}))
        self.assertIn('anest_nome_0', campos)
        self.assertIn('cat_modo_1', campos)
        self.assertIn('cat_fellow_1', campos)

    def test_ok_quando_confirmado(self):
        from repasses import views
        post = {'anest_nome_0': '__sem__', 'cat_modo_1': 'avista', 'cat_fellow_1': '__sem__'}
        self.assertEqual(views._campos_pendentes(self._rel(), post), [])


class IndexSinteticasTests(SimpleTestCase):
    """idx ÚNICO p/ linhas sintéticas (preceptoria/fellow) — sem isso elas ficavam com
    idx=0 e colidiam com a 1a linha real (hon_0 duplicado corrompia o honorário no Salvar)."""

    def test_detecta_medico_fora_do_cadastro(self):
        from types import SimpleNamespace
        from repasses import views
        from repasses.services.regras import LivroRegras, Medico as MedicoR
        livro = LivroRegras(medicos=[MedicoR(nome='Dr. Heric Sakamoto', categoria='medico')])
        rel = SimpleNamespace(blocos=[medplus.BlocoMedico(profissional='Dr. Heric Sakamoto'),
                                      medplus.BlocoMedico(profissional='Dra. Fulana Nova Silva')])
        novos = views._medicos_desconhecidos(rel, livro)
        self.assertIn('Dra. Fulana Nova Silva', novos)     # fora do cadastro
        self.assertNotIn('Dr. Heric Sakamoto', novos)      # já cadastrado

    def test_idx_unico_sem_colisao(self):
        from types import SimpleNamespace
        from repasses import views

        def proc(nome, idx=0, sintetica=False):
            p = medplus.Procedimento(None, '', '', nome, '', 1, None, None)
            p.idx, p.sintetica = idx, sintetica
            return p

        b1 = medplus.BlocoMedico(profissional='Dr. A')
        b1.procedimentos = [proc('X', 0), proc('Y', 1), proc('Preceptoria', 0, sintetica=True)]
        b2 = medplus.BlocoMedico(profissional='Dr. B (fellow)')
        b2.procedimentos = [proc('Z', 2), proc('Participação fellow', 0, sintetica=True)]
        rel = SimpleNamespace(blocos=[b1, b2])

        views._indexar_sinteticas(rel)
        idxs = [p.idx for b in rel.blocos for p in b.procedimentos]
        self.assertEqual(len(idxs), len(set(idxs)), 'nenhum idx pode repetir (senão hon_ colide)')
        self.assertEqual([b1.procedimentos[0].idx, b1.procedimentos[1].idx], [0, 1])  # reais intactos
        # sintéticas receberam idx acima do maior real (2)
        self.assertGreater(b1.procedimentos[2].idx, 2)
        self.assertGreater(b2.procedimentos[1].idx, 2)


@unittest.skipUnless(_arquivos_presentes(F22), 'amostra 22/06 ausente')
class LoteHistoricoTests(TestCase):
    """Histórico: reabrir/editar/excluir um lote; fixado (pago) bloqueia."""

    TOKEN = 'a1b2c3d4e5f6a7b8c9d0e1f2a3b4c5d6.xls'

    def _cria_lote(self):
        from repasses.models import Lote, Repasse
        with open(F22, 'rb') as f:
            conteudo = f.read()
        lote = Lote.objects.create(token=self.TOKEN, arquivo_nome='medplus_22_06.xls',
                                   unidade='Maringá', upload_conteudo=conteudo, edicoes={})
        Repasse.objects.create(lote=lote, medico='Dr. Heric Sakamoto',
                               clinica='Maringá - Matriz', valor=100, status='gerado')
        return lote

    def test_reabrir_renderiza_revisao(self):
        from repasses.models import RepasseRascunho
        lote = self._cria_lote()
        r = self.client.post(f'/lotes/{lote.id}/reabrir/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Editando o lote')
        self.assertContains(r, self.TOKEN)            # form de revisão com o token do lote
        self.assertTrue(RepasseRascunho.objects.filter(token=self.TOKEN).exists())

    def test_pago_bloqueia_reabrir_e_excluir(self):
        from repasses.models import Lote
        lote = self._cria_lote()
        rp = lote.repasses.first(); rp.status = 'pago'; rp.save()
        self.assertTrue(lote.fixado)
        r = self.client.post(f'/lotes/{lote.id}/reabrir/', follow=True)
        self.assertContains(r, 'fixado')
        self.client.post(f'/lotes/{lote.id}/excluir/')
        self.assertTrue(Lote.objects.filter(id=lote.id).exists())   # não excluiu

    def test_excluir_remove_em_cascata(self):
        from repasses.models import Lote, Repasse
        lote = self._cria_lote()
        r = self.client.post(f'/lotes/{lote.id}/excluir/', follow=True)
        self.assertFalse(Lote.objects.filter(id=lote.id).exists())
        self.assertFalse(Repasse.objects.filter(lote_id=lote.id).exists())

    def test_cadastrar_medico_novo(self):
        from repasses.models import Medico
        lote = self._cria_lote()   # garante o upload disponível (Lote.upload_conteudo)
        nome = 'Dr. Freelancer Teste da Silva'
        # sem categoria escolhida -> NÃO cadastra (o sistema não assume)
        self.client.post('/importar/cadastrar-medicos/', {
            'token': lote.token, 'novo_count': '1',
            'novo_nome_0': nome, 'novo_categoria_0': ''})
        self.assertFalse(Medico.objects.filter(nome=nome).exists())
        # com categoria + papel -> cadastra
        self.client.post('/importar/cadastrar-medicos/', {
            'token': lote.token, 'novo_count': '1', 'novo_nome_0': nome,
            'novo_categoria_0': 'medico', 'novo_razao_0': 'FREELANCER LTDA',
            'novo_regra_0': 'R$ 100/consulta', 'novo_papeis_0': 'anestesista'})
        m = Medico.objects.filter(nome=nome).first()
        self.assertIsNotNone(m)
        self.assertEqual(m.categoria, 'medico')
        self.assertEqual(m.razao_social, 'FREELANCER LTDA')
        self.assertTrue(m.eh_anestesista)

    def test_baixar_tudo_zip(self):
        import datetime
        import zipfile
        from repasses.models import ArquivoSaida
        lote = self._cria_lote()
        lote.periodo_inicio = datetime.date(2026, 6, 27)
        lote.periodo_fim = datetime.date(2026, 6, 29)
        lote.save(update_fields=['periodo_inicio', 'periodo_fim'])
        ArquivoSaida.objects.create(lote=lote, grupo='Importação OMIE',
                                    nome='OMIE_Contas_a_Pagar_27-29 de Junho.xlsx', conteudo=b'PK-fake-1')
        ArquivoSaida.objects.create(lote=lote, grupo='Repasse — Dr. Heric',
                                    nome='Dr. Heric 22-06.pdf', conteudo=b'PDF-fake-2')
        ArquivoSaida.objects.create(lote=lote, grupo='Repasse — Dr. Heric',
                                    nome='Dr. Heric 22-06.xlsx', conteudo=b'XLSX-fake-3')
        r = self.client.get(f'/lotes/{lote.id}/baixar-zip/')
        self.assertEqual(r.status_code, 200)
        # nome do zip: "Repasses_Médicos_27-29.zip" (acento via filename*=UTF-8'')
        disp = r['Content-Disposition']
        self.assertIn('Repasses_M', disp)
        self.assertIn('27-29', disp)
        conteudo = b''.join(r.streaming_content)
        with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
            nomes = zf.namelist()
        self.assertEqual(len(nomes), 3)
        # agrupado por TIPO (OMIE / PDF / XLSX), não por médico
        self.assertTrue(any(n.startswith('OMIE/') for n in nomes))
        self.assertTrue(any(n.startswith('PDF/') and n.endswith('.pdf') for n in nomes))
        self.assertTrue(any(n.startswith('XLSX/') and n.endswith('.xlsx') for n in nomes))

    def test_salvar_mostra_previa(self):
        # Após Salvar, a revisão mostra a prévia dos arquivos que serão gerados.
        lote = self._cria_lote()
        self.client.post(f'/lotes/{lote.id}/reabrir/')      # restaura o upload no disco
        r = self.client.post('/importar/salvar/', {'token': self.TOKEN})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Arquivos que serão gerados')
        self.assertContains(r, 'OMIE_Contas_a_Pagar')
